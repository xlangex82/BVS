import os, uuid
import shutil
import tempfile
import zipfile
import calendar
import io
import json
import logging
import re
from urllib.parse import urlparse, unquote
from datetime import datetime, date, timedelta
from pathlib import Path
from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory, jsonify, send_file, current_app
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import or_, func
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, BaseDocTemplate, PageTemplate, Frame, FrameBreak, Paragraph, Spacer, Table, LongTable, TableStyle, KeepTogether, Image as RLImage
from xml.sax.saxutils import escape
from pypdf import PdfReader, PdfWriter

db = SQLAlchemy()
DEFAULT_STATUSES = ["Entwurf","Beworben","Eingangsbestätigung","Telefoninterview","Vorstellungsgespräch","Zweitgespräch","Angebot","Zusage","Absage","Zurückgezogen"]
ALLOWED_EXTENSIONS = {"pdf","doc","docx","odt","txt","jpg","jpeg","png","webp","xls","xlsx","csv","zip"}

LOGO_EXTENSIONS = {"png","jpg","jpeg","webp"}
DEFAULT_APP_SETTINGS = {
    "app_name": "B-V-S",
    "app_subname": '"B"ewerbungs"V"erwaltungs"S"ystem',
    "author": "Peter Lange",
    "copyright_holder": "Lange-IT.com",
    "website": "https://lange-it.com",
    "footer_text": "",
    "logo_filename": "",
    "overview_default_mode": "active",
}

class Status(db.Model):
    __tablename__="statuses"
    id=db.Column(db.Integer,primary_key=True); name=db.Column(db.String(100),unique=True,nullable=False); sort_order=db.Column(db.Integer,nullable=False,default=0); created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow)
    applications=db.relationship("Application",back_populates="status")

class Application(db.Model):
    __tablename__="applications"
    id=db.Column(db.Integer,primary_key=True); reference=db.Column(db.String(30),unique=True,nullable=False); company=db.Column(db.String(200),nullable=False); contact_person=db.Column(db.String(200)); position=db.Column(db.String(200),nullable=False); location=db.Column(db.String(200)); application_date=db.Column(db.Date); status_id=db.Column(db.Integer,db.ForeignKey("statuses.id"),nullable=False); next_action=db.Column(db.String(255)); follow_up_date=db.Column(db.Date); job_url=db.Column(db.String(1000)); email=db.Column(db.String(255)); phone=db.Column(db.String(100)); salary_expectation=db.Column(db.String(100)); notes=db.Column(db.Text); created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow); updated_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow,onupdate=datetime.utcnow)
    status=db.relationship("Status",back_populates="applications")
    history=db.relationship("History",back_populates="application",cascade="all, delete-orphan",order_by="History.created_at.desc()")
    attachments=db.relationship("Attachment",back_populates="application",cascade="all, delete-orphan")
    details=db.relationship("ApplicationDetail",back_populates="application",uselist=False,cascade="all, delete-orphan")
    cvs=db.relationship("ApplicationCV",back_populates="application",cascade="all, delete-orphan",order_by="ApplicationCV.created_at.desc()")
    company_details=db.relationship("ApplicationCompanyDetail",back_populates="application",uselist=False,cascade="all, delete-orphan")
    letters=db.relationship("ApplicationLetter",back_populates="application",cascade="all, delete-orphan",order_by="ApplicationLetter.created_at.desc()")

class ApplicationDetail(db.Model):
    __tablename__="application_details"
    id=db.Column(db.Integer,primary_key=True); application_id=db.Column(db.Integer,db.ForeignKey("applications.id"),unique=True,nullable=False)
    source=db.Column(db.String(150)); employment_type=db.Column(db.String(100)); work_model=db.Column(db.String(100)); priority=db.Column(db.String(30),default="Normal")
    interview_date=db.Column(db.Date); interview_location=db.Column(db.String(255)); offered_salary=db.Column(db.String(100)); rejection_reason=db.Column(db.String(500))
    reminder_enabled=db.Column(db.Boolean,nullable=False,default=True); reminder_days_before=db.Column(db.Integer,nullable=False,default=2)
    application=db.relationship("Application",back_populates="details")



class StatusClassification(db.Model):
    __tablename__="status_classifications"
    id=db.Column(db.Integer,primary_key=True)
    status_id=db.Column(db.Integer,db.ForeignKey("statuses.id"),nullable=False,unique=True,index=True)
    is_completed=db.Column(db.Boolean,nullable=False,default=False)
    created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow)
    updated_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow,onupdate=datetime.utcnow)
    status=db.relationship("Status",backref=db.backref("classification",uselist=False,cascade="all, delete-orphan"))

class ApplicationCompanyDetail(db.Model):
    __tablename__="application_company_details"
    id=db.Column(db.Integer,primary_key=True)
    application_id=db.Column(db.Integer,db.ForeignKey("applications.id"),unique=True,nullable=False,index=True)
    department=db.Column(db.String(200)); street=db.Column(db.String(255)); postal_code=db.Column(db.String(30)); city=db.Column(db.String(160)); country=db.Column(db.String(120))
    application=db.relationship("Application",back_populates="company_details")

class CoverLetterMaster(db.Model):
    __tablename__="cover_letter_master"
    id=db.Column(db.Integer,primary_key=True)
    subject=db.Column(db.String(500),nullable=False,default="Bewerbung als {{position}}")
    salutation=db.Column(db.String(500),nullable=False,default="Sehr geehrte Damen und Herren,")
    intro_text=db.Column(db.Text)
    body_text=db.Column(db.Text)
    motivation_text=db.Column(db.Text)
    salary_text=db.Column(db.Text)
    closing_text=db.Column(db.Text)
    signoff=db.Column(db.String(255),nullable=False,default="Mit freundlichen Grüßen")
    created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow); updated_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow,onupdate=datetime.utcnow)

class CoverLetterTemplate(db.Model):
    __tablename__="cover_letter_templates"
    id=db.Column(db.Integer,primary_key=True)
    name=db.Column(db.String(120),nullable=False); slug=db.Column(db.String(80),unique=True,nullable=False); description=db.Column(db.String(500))
    accent_color=db.Column(db.String(20),nullable=False,default="#172033"); font_family=db.Column(db.String(40),nullable=False,default="Helvetica"); font_scale=db.Column(db.Float,nullable=False,default=1.0); page_margin_mm=db.Column(db.Integer,nullable=False,default=20)
    show_logo=db.Column(db.Boolean,nullable=False,default=False); is_active=db.Column(db.Boolean,nullable=False,default=True); sort_order=db.Column(db.Integer,nullable=False,default=0)
    created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow); updated_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow,onupdate=datetime.utcnow)
    letters=db.relationship("ApplicationLetter",back_populates="template")

class CoverLetterTemplateCover(db.Model):
    __tablename__="cover_letter_template_covers"
    id=db.Column(db.Integer,primary_key=True)
    template_id=db.Column(db.Integer,db.ForeignKey("cover_letter_templates.id"),nullable=False,unique=True,index=True)
    enabled=db.Column(db.Boolean,nullable=False,default=False)
    image_filename=db.Column(db.String(500))
    fit_mode=db.Column(db.String(30),nullable=False,default="cover")
    background_color=db.Column(db.String(20),nullable=False,default="#FFFFFF")
    created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow)
    updated_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow,onupdate=datetime.utcnow)
    template=db.relationship("CoverLetterTemplate",backref=db.backref("cover_style",uselist=False,cascade="all, delete-orphan"))

class CoverLetterCoverBlock(db.Model):
    __tablename__="cover_letter_cover_blocks"
    id=db.Column(db.Integer,primary_key=True)
    template_id=db.Column(db.Integer,db.ForeignKey("cover_letter_templates.id"),nullable=False,index=True)
    block_key=db.Column(db.String(80),nullable=False)
    label=db.Column(db.String(120),nullable=False)
    content_template=db.Column(db.Text,nullable=False,default="")
    x_mm=db.Column(db.Float,nullable=False,default=15.0)
    y_mm=db.Column(db.Float,nullable=False,default=15.0)
    width_mm=db.Column(db.Float,nullable=False,default=80.0)
    font_family=db.Column(db.String(40),nullable=False,default="Helvetica")
    font_size=db.Column(db.Float,nullable=False,default=12.0)
    text_color=db.Column(db.String(20),nullable=False,default="#FFFFFF")
    align=db.Column(db.String(20),nullable=False,default="left")
    bold=db.Column(db.Boolean,nullable=False,default=False)
    italic=db.Column(db.Boolean,nullable=False,default=False)
    enabled=db.Column(db.Boolean,nullable=False,default=True)
    sort_order=db.Column(db.Integer,nullable=False,default=0)
    created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow)
    updated_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow,onupdate=datetime.utcnow)
    template=db.relationship("CoverLetterTemplate",backref=db.backref("cover_blocks",cascade="all, delete-orphan",order_by="CoverLetterCoverBlock.sort_order"))

class ApplicationLetterCoverSnapshot(db.Model):
    __tablename__="application_letter_cover_snapshots"
    id=db.Column(db.Integer,primary_key=True)
    application_letter_id=db.Column(db.Integer,db.ForeignKey("application_letters.id"),nullable=False,unique=True,index=True)
    data_json=db.Column(db.Text,nullable=False,default="{}")
    created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow)
    letter=db.relationship("ApplicationLetter",backref=db.backref("cover_snapshot",uselist=False,cascade="all, delete-orphan"))

class CoverLetterFooterStyle(db.Model):
    __tablename__="cover_letter_footer_styles"
    id=db.Column(db.Integer,primary_key=True)
    template_id=db.Column(db.Integer,db.ForeignKey("cover_letter_templates.id"),nullable=False,unique=True,index=True)
    enabled=db.Column(db.Boolean,nullable=False,default=False)
    layout=db.Column(db.String(40),nullable=False,default="centered")
    show_divider=db.Column(db.Boolean,nullable=False,default=True)
    divider_color=db.Column(db.String(20),nullable=False,default="#D0D5DD")
    divider_width_pct=db.Column(db.Integer,nullable=False,default=90)
    logo_height_mm=db.Column(db.Float,nullable=False,default=14.0)
    logo_gap_mm=db.Column(db.Float,nullable=False,default=16.0)
    primary_color=db.Column(db.String(20),nullable=False,default="#5B6E6A")
    secondary_color=db.Column(db.String(20),nullable=False,default="#E5E7EB")
    created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow)
    updated_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow,onupdate=datetime.utcnow)
    template=db.relationship("CoverLetterTemplate",backref=db.backref("footer_style",uselist=False,cascade="all, delete-orphan"))

class CertificationLogo(db.Model):
    __tablename__="certification_logos"
    id=db.Column(db.Integer,primary_key=True)
    name=db.Column(db.String(200),nullable=False)
    issuer=db.Column(db.String(200))
    filename=db.Column(db.String(500),nullable=False)
    sort_order=db.Column(db.Integer,nullable=False,default=0)
    is_active=db.Column(db.Boolean,nullable=False,default=True)
    created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow)
    updated_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow,onupdate=datetime.utcnow)

class CoverLetterTemplateCertification(db.Model):
    __tablename__="cover_letter_template_certifications"
    id=db.Column(db.Integer,primary_key=True)
    template_id=db.Column(db.Integer,db.ForeignKey("cover_letter_templates.id"),nullable=False,index=True)
    certification_id=db.Column(db.Integer,db.ForeignKey("certification_logos.id"),nullable=False,index=True)
    enabled=db.Column(db.Boolean,nullable=False,default=True)
    sort_order=db.Column(db.Integer,nullable=False,default=0)
    template=db.relationship("CoverLetterTemplate",backref=db.backref("footer_certification_links",cascade="all, delete-orphan"))
    certification=db.relationship("CertificationLogo")
    __table_args__=(db.UniqueConstraint("template_id","certification_id",name="uq_letter_template_certification"),)

class ApplicationLetter(db.Model):
    __tablename__="application_letters"
    id=db.Column(db.Integer,primary_key=True)
    application_id=db.Column(db.Integer,db.ForeignKey("applications.id"),nullable=False,index=True); template_id=db.Column(db.Integer,db.ForeignKey("cover_letter_templates.id"),nullable=False)
    title=db.Column(db.String(200),nullable=False,default="Anschreiben")
    recipient_company=db.Column(db.String(200)); recipient_department=db.Column(db.String(200)); recipient_contact=db.Column(db.String(200)); recipient_street=db.Column(db.String(255)); recipient_postal_code=db.Column(db.String(30)); recipient_city=db.Column(db.String(160)); recipient_country=db.Column(db.String(120))
    sender_name=db.Column(db.String(255)); sender_address=db.Column(db.String(255)); sender_postal_code=db.Column(db.String(30)); sender_city=db.Column(db.String(160)); sender_email=db.Column(db.String(255)); sender_phone=db.Column(db.String(100))
    letter_date=db.Column(db.Date); subject=db.Column(db.String(500)); salutation=db.Column(db.String(500)); intro_text=db.Column(db.Text); body_text=db.Column(db.Text); motivation_text=db.Column(db.Text); salary_text=db.Column(db.Text); closing_text=db.Column(db.Text); signoff=db.Column(db.String(255)); signature_filename=db.Column(db.String(500))
    created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow); updated_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow,onupdate=datetime.utcnow)
    application=db.relationship("Application",back_populates="letters"); template=db.relationship("CoverLetterTemplate",back_populates="letters")

class CVProfile(db.Model):
    __tablename__="cv_profile"
    id=db.Column(db.Integer,primary_key=True)
    first_name=db.Column(db.String(120)); last_name=db.Column(db.String(120)); professional_title=db.Column(db.String(200))
    birth_date=db.Column(db.Date); birth_place=db.Column(db.String(200)); address=db.Column(db.String(255)); postal_code=db.Column(db.String(30)); city=db.Column(db.String(120)); country=db.Column(db.String(120))
    email=db.Column(db.String(255)); phone=db.Column(db.String(100)); website=db.Column(db.String(500)); linkedin=db.Column(db.String(500)); xing=db.Column(db.String(500)); github=db.Column(db.String(500)); summary=db.Column(db.Text)
    created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow); updated_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow,onupdate=datetime.utcnow)

class CVExperience(db.Model):
    __tablename__="cv_experience"
    id=db.Column(db.Integer,primary_key=True); employer=db.Column(db.String(200),nullable=False); position=db.Column(db.String(200),nullable=False); location=db.Column(db.String(200)); start_date=db.Column(db.Date); end_date=db.Column(db.Date); is_current=db.Column(db.Boolean,nullable=False,default=False); description=db.Column(db.Text); sort_order=db.Column(db.Integer,nullable=False,default=0); created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow); updated_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow,onupdate=datetime.utcnow)

class CVEducation(db.Model):
    __tablename__="cv_education"
    id=db.Column(db.Integer,primary_key=True); institution=db.Column(db.String(200),nullable=False); degree=db.Column(db.String(200)); field_of_study=db.Column(db.String(200)); location=db.Column(db.String(200)); start_date=db.Column(db.Date); end_date=db.Column(db.Date); is_current=db.Column(db.Boolean,nullable=False,default=False); description=db.Column(db.Text); sort_order=db.Column(db.Integer,nullable=False,default=0); created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow); updated_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow,onupdate=datetime.utcnow)

class CVSkill(db.Model):
    __tablename__="cv_skills"
    id=db.Column(db.Integer,primary_key=True); name=db.Column(db.String(160),nullable=False); category=db.Column(db.String(120)); level=db.Column(db.String(80)); sort_order=db.Column(db.Integer,nullable=False,default=0); created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow); updated_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow,onupdate=datetime.utcnow)

class CVLanguage(db.Model):
    __tablename__="cv_languages"
    id=db.Column(db.Integer,primary_key=True); language=db.Column(db.String(120),nullable=False); level=db.Column(db.String(120)); sort_order=db.Column(db.Integer,nullable=False,default=0); created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow); updated_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow,onupdate=datetime.utcnow)

class CVCertification(db.Model):
    __tablename__="cv_certifications"
    id=db.Column(db.Integer,primary_key=True); name=db.Column(db.String(200),nullable=False); issuer=db.Column(db.String(200)); issue_date=db.Column(db.Date); expiry_date=db.Column(db.Date); credential_id=db.Column(db.String(200)); credential_url=db.Column(db.String(500)); description=db.Column(db.Text); sort_order=db.Column(db.Integer,nullable=False,default=0); created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow); updated_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow,onupdate=datetime.utcnow)

class CVProject(db.Model):
    __tablename__="cv_projects"
    id=db.Column(db.Integer,primary_key=True); name=db.Column(db.String(200),nullable=False); role=db.Column(db.String(200)); start_date=db.Column(db.Date); end_date=db.Column(db.Date); url=db.Column(db.String(500)); technologies=db.Column(db.String(500)); description=db.Column(db.Text); sort_order=db.Column(db.Integer,nullable=False,default=0); created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow); updated_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow,onupdate=datetime.utcnow)

class CVOther(db.Model):
    __tablename__="cv_other"
    id=db.Column(db.Integer,primary_key=True); category=db.Column(db.String(120)); title=db.Column(db.String(200),nullable=False); organization=db.Column(db.String(200)); location=db.Column(db.String(200)); start_date=db.Column(db.Date); end_date=db.Column(db.Date); description=db.Column(db.Text); sort_order=db.Column(db.Integer,nullable=False,default=0); created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow); updated_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow,onupdate=datetime.utcnow)


class CVTemplate(db.Model):
    __tablename__="cv_templates"
    id=db.Column(db.Integer,primary_key=True)
    name=db.Column(db.String(120),nullable=False)
    slug=db.Column(db.String(80),unique=True,nullable=False)
    description=db.Column(db.String(500))
    sort_order=db.Column(db.Integer,nullable=False,default=0)
    is_active=db.Column(db.Boolean,nullable=False,default=True)
    created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow)
    updated_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow,onupdate=datetime.utcnow)
    application_cvs=db.relationship("ApplicationCV",back_populates="template")

class CVTemplateStyle(db.Model):
    __tablename__="cv_template_styles"
    id=db.Column(db.Integer,primary_key=True)
    template_id=db.Column(db.Integer,db.ForeignKey("cv_templates.id"),nullable=False,unique=True,index=True)
    base_layout=db.Column(db.String(30),nullable=False,default="classic")
    accent_color=db.Column(db.String(20),nullable=False,default="#172033")
    font_family=db.Column(db.String(40),nullable=False,default="Helvetica")
    font_scale=db.Column(db.Float,nullable=False,default=1.0)
    page_margin_mm=db.Column(db.Integer,nullable=False,default=15)
    show_logo=db.Column(db.Boolean,nullable=False,default=True)
    show_company=db.Column(db.Boolean,nullable=False,default=True)
    show_target_position=db.Column(db.Boolean,nullable=False,default=True)
    section_order=db.Column(db.String(500),nullable=False,default="experience,education,projects,skills,languages,certifications,other")
    created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow)
    updated_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow,onupdate=datetime.utcnow)
    template=db.relationship("CVTemplate",backref=db.backref("style",uselist=False,cascade="all, delete-orphan"))

class ApplicationCV(db.Model):
    __tablename__="application_cvs"
    id=db.Column(db.Integer,primary_key=True)
    application_id=db.Column(db.Integer,db.ForeignKey("applications.id"),nullable=False,index=True)
    template_id=db.Column(db.Integer,db.ForeignKey("cv_templates.id"),nullable=False)
    title=db.Column(db.String(200),nullable=False,default="Lebenslauf")
    target_company=db.Column(db.String(200))
    target_position=db.Column(db.String(200))
    profile_summary=db.Column(db.Text)
    profile_json=db.Column(db.Text,nullable=False,default="{}")
    created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow)
    updated_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow,onupdate=datetime.utcnow)
    application=db.relationship("Application",back_populates="cvs")
    template=db.relationship("CVTemplate",back_populates="application_cvs")
    entries=db.relationship("ApplicationCVEntry",back_populates="application_cv",cascade="all, delete-orphan",order_by="ApplicationCVEntry.section, ApplicationCVEntry.sort_order, ApplicationCVEntry.id")

class ApplicationCVHeader(db.Model):
    __tablename__="application_cv_headers"
    id=db.Column(db.Integer,primary_key=True)
    application_cv_id=db.Column(db.Integer,db.ForeignKey("application_cvs.id"),nullable=False,unique=True,index=True)
    document_title=db.Column(db.String(160),nullable=False,default="Lebenslauf")
    header_subtitle=db.Column(db.String(240))
    header_layout=db.Column(db.String(40),nullable=False,default="photo_right")
    show_document_title=db.Column(db.Boolean,nullable=False,default=True)
    show_professional_title=db.Column(db.Boolean,nullable=False,default=True)
    show_target_company=db.Column(db.Boolean,nullable=False,default=True)
    show_target_position=db.Column(db.Boolean,nullable=False,default=True)
    show_contact=db.Column(db.Boolean,nullable=False,default=True)
    logo_filename=db.Column(db.String(500))
    photo_filename=db.Column(db.String(500))
    created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow)
    updated_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow,onupdate=datetime.utcnow)
    application_cv=db.relationship("ApplicationCV",backref=db.backref("header",uselist=False,cascade="all, delete-orphan"))

class ApplicationCVEntry(db.Model):
    __tablename__="application_cv_entries"
    id=db.Column(db.Integer,primary_key=True)
    application_cv_id=db.Column(db.Integer,db.ForeignKey("application_cvs.id"),nullable=False,index=True)
    section=db.Column(db.String(50),nullable=False,index=True)
    source_id=db.Column(db.Integer)
    sort_order=db.Column(db.Integer,nullable=False,default=0)
    visible=db.Column(db.Boolean,nullable=False,default=True)
    data_json=db.Column(db.Text,nullable=False,default="{}")
    created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow)
    updated_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow,onupdate=datetime.utcnow)
    application_cv=db.relationship("ApplicationCV",back_populates="entries")


class UniversalCV(db.Model):
    __tablename__="universal_cvs"
    id=db.Column(db.Integer,primary_key=True)
    template_id=db.Column(db.Integer,db.ForeignKey("cv_templates.id"),nullable=False)
    title=db.Column(db.String(200),nullable=False,default="Lebenslauf")
    target_company=db.Column(db.String(200))
    target_position=db.Column(db.String(200))
    profile_summary=db.Column(db.Text)
    profile_json=db.Column(db.Text,nullable=False,default="{}")
    created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow)
    updated_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow,onupdate=datetime.utcnow)
    template=db.relationship("CVTemplate")
    entries=db.relationship("UniversalCVEntry",back_populates="universal_cv",cascade="all, delete-orphan",order_by="UniversalCVEntry.section, UniversalCVEntry.sort_order, UniversalCVEntry.id")

class UniversalCVHeader(db.Model):
    __tablename__="universal_cv_headers"
    id=db.Column(db.Integer,primary_key=True)
    universal_cv_id=db.Column(db.Integer,db.ForeignKey("universal_cvs.id"),nullable=False,unique=True,index=True)
    document_title=db.Column(db.String(160),nullable=False,default="Lebenslauf")
    header_subtitle=db.Column(db.String(240))
    header_layout=db.Column(db.String(40),nullable=False,default="photo_right")
    show_document_title=db.Column(db.Boolean,nullable=False,default=True)
    show_professional_title=db.Column(db.Boolean,nullable=False,default=True)
    show_target_company=db.Column(db.Boolean,nullable=False,default=False)
    show_target_position=db.Column(db.Boolean,nullable=False,default=True)
    show_contact=db.Column(db.Boolean,nullable=False,default=True)
    logo_filename=db.Column(db.String(500))
    photo_filename=db.Column(db.String(500))
    created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow)
    updated_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow,onupdate=datetime.utcnow)
    universal_cv=db.relationship("UniversalCV",backref=db.backref("header",uselist=False,cascade="all, delete-orphan"))

class UniversalCVEntry(db.Model):
    __tablename__="universal_cv_entries"
    id=db.Column(db.Integer,primary_key=True)
    universal_cv_id=db.Column(db.Integer,db.ForeignKey("universal_cvs.id"),nullable=False,index=True)
    section=db.Column(db.String(50),nullable=False,index=True)
    source_id=db.Column(db.Integer)
    sort_order=db.Column(db.Integer,nullable=False,default=0)
    visible=db.Column(db.Boolean,nullable=False,default=True)
    data_json=db.Column(db.Text,nullable=False,default="{}")
    created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow)
    updated_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow,onupdate=datetime.utcnow)
    universal_cv=db.relationship("UniversalCV",back_populates="entries")

class History(db.Model):
    __tablename__="history"
    id=db.Column(db.Integer,primary_key=True); application_id=db.Column(db.Integer,db.ForeignKey("applications.id"),nullable=False); event_type=db.Column(db.String(50),nullable=False); message=db.Column(db.String(1000),nullable=False); created_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow)
    application=db.relationship("Application",back_populates="history")

class Attachment(db.Model):
    __tablename__="attachments"
    id=db.Column(db.Integer,primary_key=True); application_id=db.Column(db.Integer,db.ForeignKey("applications.id"),nullable=False); original_name=db.Column(db.String(255),nullable=False); stored_name=db.Column(db.String(255),unique=True,nullable=False); uploaded_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow)
    application=db.relationship("Application",back_populates="attachments")

class AppSetting(db.Model):
    __tablename__="app_settings"
    id=db.Column(db.Integer,primary_key=True)
    key=db.Column(db.String(100),unique=True,nullable=False)
    value=db.Column(db.Text)
    updated_at=db.Column(db.DateTime,nullable=False,default=datetime.utcnow,onupdate=datetime.utcnow)


def ensure_default_settings():
    changed=False
    for key, value in DEFAULT_APP_SETTINGS.items():
        if not AppSetting.query.filter_by(key=key).first():
            db.session.add(AppSetting(key=key,value=value))
            changed=True
    if changed:
        db.session.commit()

def get_app_settings():
    values=dict(DEFAULT_APP_SETTINGS)
    for item in AppSetting.query.all():
        values[item.key]=item.value or ""
    return values

def set_app_setting(key, value):
    item=AppSetting.query.filter_by(key=key).first()
    if item is None:
        item=AppSetting(key=key)
        db.session.add(item)
    item.value=value or ""

def logo_path():
    settings=get_app_settings()
    filename=settings.get("logo_filename","")
    if not filename:
        return None
    path=Path(current_app.config["UPLOAD_DIR"]) / "branding" / filename
    return path if path.exists() else None

def application_filter_query():
    q=request.args.get("q","").strip()
    sid=request.args.get("status",type=int)
    sort=request.args.get("sort","follow_up")
    date_from_raw=request.args.get("date_from","").strip()
    date_to_raw=request.args.get("date_to","").strip()
    date_from=parse_date(date_from_raw) if date_from_raw else None
    date_to=parse_date(date_to_raw) if date_to_raw else None

    process_mode=request.args.get("process_mode","").strip()
    if process_mode not in ("all","active","completed"):
        process_mode=get_app_settings().get("overview_default_mode","active")
    if process_mode not in ("all","active","completed"):
        process_mode="active"

    query=Application.query
    if q:
        term=f"%{q}%"
        query=query.filter(or_(
            Application.reference.ilike(term),
            Application.company.ilike(term),
            Application.position.ilike(term),
            Application.contact_person.ilike(term),
            Application.location.ilike(term)
        ))
    if sid:
        query=query.filter(Application.status_id==sid)
    if date_from:
        query=query.filter(Application.application_date >= date_from)
    if date_to:
        query=query.filter(Application.application_date <= date_to)

    completed_ids=completed_status_ids()
    if process_mode=="active" and completed_ids:
        query=query.filter(~Application.status_id.in_(completed_ids))
    elif process_mode=="completed":
        if completed_ids:
            query=query.filter(Application.status_id.in_(completed_ids))
        else:
            query=query.filter(db.false())

    if sort=="company":
        query=query.order_by(Application.company.asc(),Application.position.asc())
    elif sort=="created":
        query=query.order_by(Application.created_at.desc())
    elif sort=="application_date":
        query=query.order_by(Application.application_date.desc(),Application.company.asc())
    else:
        query=query.order_by(
            Application.follow_up_date.is_(None),
            Application.follow_up_date.asc(),
            Application.updated_at.desc()
        )
    return query, {
        "q": q,
        "status": sid,
        "sort": sort,
        "date_from": date_from_raw,
        "date_to": date_to_raw,
        "process_mode": process_mode,
    }


def allowed_file(f): return "." in f and f.rsplit(".",1)[1].lower() in ALLOWED_EXTENSIONS
def parse_date(v): return datetime.strptime(v,"%Y-%m-%d").date() if v else None
def get_statuses(): return Status.query.order_by(Status.sort_order,Status.name).all()
def log_event(a,t,m): db.session.add(History(application=a,event_type=t,message=m))
def make_reference():
    y=datetime.now().year; n=Application.query.filter(Application.reference.like(f"{y}-%")).count()+1; r=f"{y}-{n:03d}"
    while Application.query.filter_by(reference=r).first(): n+=1; r=f"{y}-{n:03d}"
    return r

def detail_for(a):
    if a.details is None:
        a.details=ApplicationDetail(reminder_enabled=True,reminder_days_before=2,priority="Normal")
    return a.details

def reminder_items():
    items=[]; today=date.today()
    apps=Application.query.join(Status).filter(~Status.name.in_(["Absage","Zurückgezogen"]),Application.follow_up_date.is_not(None)).all()
    for a in apps:
        d=a.details; enabled=True if d is None else d.reminder_enabled; days=2 if d is None else max(0,d.reminder_days_before or 0)
        if enabled and a.follow_up_date-timedelta(days=days) <= today:
            items.append((a,(a.follow_up_date-today).days))
    return sorted(items,key=lambda x:x[0].follow_up_date)



def pdf_text(value):
    if value is None or value == "":
        return "-"
    return escape(str(value)).replace("\n", "<br/>")

def format_date_de(value):
    return value.strftime("%d.%m.%Y") if value else "-"

def build_application_pdf(application):
    """Create a compact DIN-A4 PDF summary for one application."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=16*mm,
        leftMargin=16*mm,
        topMargin=15*mm,
        bottomMargin=15*mm,
        title=f"Bewerbung {application.reference}",
        author=get_app_settings().get("author","Peter Lange"),
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="DocTitle", parent=styles["Title"], fontSize=18, leading=22, spaceAfter=4))
    styles.add(ParagraphStyle(name="DocSub", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#667085"), spaceAfter=10))
    styles.add(ParagraphStyle(name="Section", parent=styles["Heading2"], fontSize=11, leading=14, spaceBefore=8, spaceAfter=5, textColor=colors.HexColor("#172033")))
    styles.add(ParagraphStyle(name="CellLabel", parent=styles["Normal"], fontSize=7.5, leading=9, textColor=colors.HexColor("#667085")))
    styles.add(ParagraphStyle(name="CellValue", parent=styles["Normal"], fontSize=9.2, leading=11.5))
    styles.add(ParagraphStyle(name="Notes", parent=styles["Normal"], fontSize=9, leading=12))
    styles.add(ParagraphStyle(name="Footer", parent=styles["Normal"], fontSize=7.5, leading=9, textColor=colors.HexColor("#667085"), alignment=TA_CENTER))

    d = application.details
    settings=get_app_settings()
    story = []
    lp=logo_path()
    if lp:
        try:
            story.append(RLImage(str(lp),width=18*mm,height=18*mm,kind="proportional"))
            story.append(Spacer(1,2*mm))
        except Exception:
            pass
    story += [
        Paragraph(pdf_text(application.company), styles["DocTitle"]),
        Paragraph(f"{pdf_text(application.position)} &nbsp;&nbsp;|&nbsp;&nbsp; {pdf_text(application.reference)}", styles["DocSub"]),
    ]

    def field(label, value):
        return [Paragraph(escape(label), styles["CellLabel"]), Paragraph(pdf_text(value), styles["CellValue"])]

    def section(title, rows):
        table_data=[]
        for left_label,left_val,right_label,right_val in rows:
            table_data.append([
                field(left_label,left_val), field(right_label,right_val)
            ])
        t=Table(table_data, colWidths=[88*mm,88*mm], hAlign="LEFT")
        t.setStyle(TableStyle([
            ("VALIGN",(0,0),(-1,-1),"TOP"),
            ("LEFTPADDING",(0,0),(-1,-1),0),
            ("RIGHTPADDING",(0,0),(-1,-1),6),
            ("TOPPADDING",(0,0),(-1,-1),3),
            ("BOTTOMPADDING",(0,0),(-1,-1),5),
            ("LINEBELOW",(0,0),(-1,-1),0.25,colors.HexColor("#e4e7ec")),
        ]))
        story.extend([Paragraph(title, styles["Section"]), t])

    section("Bewerbung", [
        ("Status", application.status.name if application.status else "-", "Bewerbungsdatum", format_date_de(application.application_date)),
        ("Ansprechpartner", application.contact_person, "Standort", application.location),
        ("Quelle", d.source if d else None, "Priorität", d.priority if d else None),
        ("Beschäftigungsart", d.employment_type if d else None, "Arbeitsmodell", d.work_model if d else None),
    ])
    section("Planung & Wiedervorlage", [
        ("Nächste Aktion", application.next_action, "Wiedervorlage", format_date_de(application.follow_up_date)),
        ("Erinnerung", ("Aktiv" if d and d.reminder_enabled else "Aus") if d else "Aktiv", "Erinnerung vorher", f"{d.reminder_days_before} Tag(e)" if d else "2 Tag(e)"),
        ("Gesprächsdatum", format_date_de(d.interview_date) if d else "-", "Gesprächsort / Link", d.interview_location if d else None),
    ])
    section("Kontakt & Konditionen", [
        ("E-Mail", application.email, "Telefon", application.phone),
        ("Gehaltsvorstellung", application.salary_expectation, "Angebotenes Gehalt", d.offered_salary if d else None),
        ("Stellenausschreibung", application.job_url, "Absagegrund", d.rejection_reason if d else None),
    ])

    story.append(Paragraph("Notizen", styles["Section"]))
    notes = pdf_text(application.notes)
    note_table=Table([[Paragraph(notes, styles["Notes"])]], colWidths=[176*mm])
    note_table.setStyle(TableStyle([
        ("BOX",(0,0),(-1,-1),0.5,colors.HexColor("#d0d5dd")),
        ("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#f9fafb")),
        ("LEFTPADDING",(0,0),(-1,-1),7), ("RIGHTPADDING",(0,0),(-1,-1),7),
        ("TOPPADDING",(0,0),(-1,-1),7), ("BOTTOMPADDING",(0,0),(-1,-1),7),
    ]))
    footer_text=f"{settings.get('app_name','B-V-S')} · © {date.today().year} {settings.get('copyright_holder','Lange-IT.com')} · {settings.get('author','Peter Lange')}"
    story.extend([note_table, Spacer(1,8*mm), Paragraph(footer_text, styles["Footer"])])

    doc.build(story)
    buffer.seek(0)
    return buffer


def build_overview_pdf(applications, filters):
    settings=get_app_settings()
    buffer=io.BytesIO()
    doc=SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=10*mm,
        leftMargin=10*mm,
        topMargin=10*mm,
        bottomMargin=10*mm,
        title=f"{settings.get('app_name','B-V-S')} – Bewerbungsübersicht",
        author=settings.get("author","Peter Lange"),
    )
    styles=getSampleStyleSheet()
    styles.add(ParagraphStyle(name="OverviewTitle",parent=styles["Title"],fontSize=16,leading=19,spaceAfter=2))
    styles.add(ParagraphStyle(name="OverviewSub",parent=styles["Normal"],fontSize=8.5,leading=10,textColor=colors.HexColor("#667085"),spaceAfter=7))
    styles.add(ParagraphStyle(name="OverviewCell",parent=styles["Normal"],fontSize=7.1,leading=8.5))
    styles.add(ParagraphStyle(name="OverviewHead",parent=styles["Normal"],fontSize=7.1,leading=8.5,textColor=colors.white))
    story=[]

    lp=logo_path()
    if lp:
        try:
            logo=RLImage(str(lp),width=18*mm,height=18*mm,kind="proportional")
            heading=Table([[
                logo,
                [Paragraph(escape(settings.get("app_name","B-V-S")),styles["OverviewTitle"]),
                 Paragraph(escape(settings.get("app_subname","")),styles["OverviewSub"])]
            ]],colWidths=[23*mm,245*mm])
            heading.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"MIDDLE"),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),4)]))
            story.append(heading)
        except Exception:
            story.append(Paragraph(escape(settings.get("app_name","B-V-S")),styles["OverviewTitle"]))
    else:
        story.append(Paragraph(escape(settings.get("app_name","B-V-S")),styles["OverviewTitle"]))

    filter_parts=[]
    process_labels={"active":"Nur aktive","completed":"Nur erledigte","all":"Alle"}
    if filters.get("process_mode"): filter_parts.append(f"Vorgänge: {process_labels.get(filters['process_mode'],filters['process_mode'])}")
    if filters.get("q"): filter_parts.append(f"Suche: {filters['q']}")
    if filters.get("status"):
        st=Status.query.get(filters["status"])
        if st: filter_parts.append(f"Status: {st.name}")
    if filters.get("date_from"): filter_parts.append(f"von: {filters['date_from']}")
    if filters.get("date_to"): filter_parts.append(f"bis: {filters['date_to']}")
    subtitle=f"Bewerbungsübersicht · {len(applications)} Einträge · Stand {date.today().strftime('%d.%m.%Y')}"
    if filter_parts: subtitle += " · " + " · ".join(filter_parts)
    story.append(Paragraph(escape(subtitle),styles["OverviewSub"]))

    headers=["Ref.","Unternehmen","Position","Status","Bewerbung","Wiedervorlage","Nächste Aktion"]
    data=[[Paragraph(h,styles["OverviewHead"]) for h in headers]]
    for a in applications:
        data.append([
            Paragraph(pdf_text(a.reference),styles["OverviewCell"]),
            Paragraph(pdf_text(a.company),styles["OverviewCell"]),
            Paragraph(pdf_text(a.position),styles["OverviewCell"]),
            Paragraph(pdf_text(a.status.name if a.status else "-"),styles["OverviewCell"]),
            Paragraph(pdf_text(format_date_de(a.application_date)),styles["OverviewCell"]),
            Paragraph(pdf_text(format_date_de(a.follow_up_date)),styles["OverviewCell"]),
            Paragraph(pdf_text(a.next_action),styles["OverviewCell"]),
        ])
    if len(data)==1:
        data.append([Paragraph("Keine Bewerbungen für diese Filter.",styles["OverviewCell"])] + [""]*6)

    table=Table(data,repeatRows=1,colWidths=[20*mm,48*mm,57*mm,35*mm,25*mm,27*mm,65*mm])
    table.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#172033")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#d0d5dd")),
        ("VALIGN",(0,0),(-1,-1),"TOP"),
        ("LEFTPADDING",(0,0),(-1,-1),3),
        ("RIGHTPADDING",(0,0),(-1,-1),3),
        ("TOPPADDING",(0,0),(-1,-1),3),
        ("BOTTOMPADDING",(0,0),(-1,-1),3),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#f8fafc")]),
    ]))
    story.extend([table,Spacer(1,4*mm)])
    footer=f"© {date.today().year} {settings.get('copyright_holder','Lange-IT.com')} · Autor: {settings.get('author','Peter Lange')}"
    if settings.get("footer_text"):
        footer += f" · {settings['footer_text']}"
    story.append(Paragraph(escape(footer),styles["OverviewSub"]))
    doc.build(story)
    buffer.seek(0)
    return buffer





DEFAULT_COVER_LETTER_TEMPLATES = [
    ("Klassisch","classic","Traditionelles, klares Geschäftsbrief-Layout.",10),
    ("Modern","modern","Moderner Briefkopf mit Akzent und kompakter Absenderzeile.",20),
    ("Kompakt","compact","Platzsparendes Layout für einseitige Anschreiben.",30),
]

def ensure_cover_letter_defaults():
    master=CoverLetterMaster.query.first()
    if not master:
        master=CoverLetterMaster(
            intro_text="mit großem Interesse bewerbe ich mich auf die Position {{position}} bei {{company}}.",
            body_text="Meine beruflichen Erfahrungen und Kenntnisse möchte ich gezielt in die ausgeschriebene Position einbringen. Besonders reizt mich die Möglichkeit, meine Stärken in einem neuen Umfeld einzusetzen und weiterzuentwickeln.",
            motivation_text="Gerne überzeuge ich Sie in einem persönlichen Gespräch davon, wie ich Ihr Team unterstützen kann.",
            salary_text="Meine Gehaltsvorstellung liegt bei {{salary_expectation}} brutto jährlich.",
            closing_text="Über eine Einladung zu einem persönlichen Gespräch freue ich mich sehr.",
        )
        db.session.add(master)
    for name,slug,description,order in DEFAULT_COVER_LETTER_TEMPLATES:
        if not CoverLetterTemplate.query.filter_by(slug=slug).first():
            db.session.add(CoverLetterTemplate(name=name,slug=slug,description=description,sort_order=order,accent_color={"classic":"#172033","modern":"#2f5d8a","compact":"#3d4653"}[slug]))
    db.session.commit()

def ensure_status_classifications():
    completed_defaults={"Absage","Zurückgezogen","Zusage"}
    changed=False
    for status in Status.query.all():
        if not status.classification:
            db.session.add(StatusClassification(
                status=status,
                is_completed=status.name in completed_defaults
            ))
            changed=True
    if changed:
        db.session.commit()

def completed_status_ids():
    return [row.status_id for row in StatusClassification.query.filter_by(is_completed=True).all()]

def active_application_count():
    ids=completed_status_ids()
    query=Application.query
    if ids:
        query=query.filter(~Application.status_id.in_(ids))
    return query.count()

def company_detail_for(application):
    if application.company_details: return application.company_details
    detail=ApplicationCompanyDetail(application=application)
    db.session.add(detail)
    return detail

def cover_letter_variables(application):
    profile=CVProfile.query.first()
    cd=application.company_details
    d=application.details
    values={
        "company":application.company or "", "department":cd.department if cd else "", "contact":application.contact_person or "", "position":application.position or "", "reference":application.reference or "", "location":application.location or "", "salary_expectation":application.salary_expectation or "", "source":d.source if d else "", "application_date":application.application_date.strftime("%d.%m.%Y") if application.application_date else "", "today":date.today().strftime("%d.%m.%Y"),
        "company_street":cd.street if cd else "", "company_postal_code":cd.postal_code if cd else "", "company_city":cd.city if cd else "", "company_country":cd.country if cd else "",
        "first_name":profile.first_name if profile else "", "last_name":profile.last_name if profile else "", "professional_title":profile.professional_title if profile else "", "email":profile.email if profile else "", "phone":profile.phone if profile else "", "address":profile.address if profile else "", "postal_code":profile.postal_code if profile else "", "city":profile.city if profile else "", "country":profile.country if profile else "",
    }
    return {k:(v or "") for k,v in values.items()}

def render_letter_placeholders(value, variables):
    if not value: return ""
    def repl(match): return str(variables.get(match.group(1).strip(), match.group(0)))
    return re.sub(r"\{\{\s*([a-zA-Z0-9_]+)\s*\}\}",repl,value)

def master_signature_filename():
    setting=AppSetting.query.filter_by(key="cover_letter_signature_filename").first()
    return (setting.value or "") if setting else ""

def create_letter_snapshot(application, template):
    master=CoverLetterMaster.query.first() or CoverLetterMaster()
    profile=CVProfile.query.first(); cd=application.company_details; vars=cover_letter_variables(application)
    letter=ApplicationLetter(
        application=application,template=template,title="Anschreiben",recipient_company=application.company,recipient_department=cd.department if cd else None,recipient_contact=application.contact_person,recipient_street=cd.street if cd else None,recipient_postal_code=cd.postal_code if cd else None,recipient_city=(cd.city if cd and cd.city else application.location),recipient_country=cd.country if cd else None,
        sender_name=(" ".join([profile.first_name or "",profile.last_name or ""]).strip() if profile else None),sender_address=profile.address if profile else None,sender_postal_code=profile.postal_code if profile else None,sender_city=profile.city if profile else None,sender_email=profile.email if profile else None,sender_phone=profile.phone if profile else None,letter_date=date.today(),
        subject=render_letter_placeholders(master.subject,vars),salutation=render_letter_placeholders(master.salutation,vars),intro_text=render_letter_placeholders(master.intro_text,vars),body_text=render_letter_placeholders(master.body_text,vars),motivation_text=render_letter_placeholders(master.motivation_text,vars),salary_text=render_letter_placeholders(master.salary_text,vars) if application.salary_expectation else None,closing_text=render_letter_placeholders(master.closing_text,vars),signoff=render_letter_placeholders(master.signoff,vars)
    )
    db.session.add(letter); db.session.flush()
    create_letter_cover_snapshot(letter)
    sig=master_signature_filename()
    if sig: letter.signature_filename=copy_cv_media(sig,f"letters/{letter.id}","signature")
    return letter

def letter_sender_line(letter):
    """Canonical sender header used by browser preview and PDF export."""
    postal_city=" ".join(x for x in [letter.sender_postal_code,letter.sender_city] if x)
    return " · ".join(x for x in [
        letter.sender_name,
        letter.sender_address,
        postal_city,
        letter.sender_email,
    ] if x)

def letter_paragraphs(letter):
    return [x for x in [letter.intro_text,letter.body_text,letter.motivation_text,letter.salary_text,letter.closing_text] if x and x.strip()]

DEFAULT_COVER_BLOCKS = [
    {
        "block_key":"applicant","label":"Bewerber",
        "content_template":"{{applicant_name}}\n{{applicant_address}}\n{{applicant_postal_city}}\n\nTel.: {{applicant_phone}}\nE-Mail: {{applicant_email}}",
        "x_mm":12.0,"y_mm":16.0,"width_mm":75.0,"font_size":10.5,"text_color":"#FFFFFF","align":"left","bold":True,"italic":False,"sort_order":10
    },
    {
        "block_key":"application","label":"Bewerbung",
        "content_template":"B E W E R B U N G\n\nals\n\n{{position}}\n\nvia {{source}}",
        "x_mm":64.0,"y_mm":102.0,"width_mm":100.0,"font_size":17.0,"text_color":"#FFFFFF","align":"center","bold":True,"italic":True,"sort_order":20
    },
    {
        "block_key":"recipient","label":"Empfänger",
        "content_template":"{{company}}\n{{department}}\n{{company_street}}\n{{company_postal_city}}",
        "x_mm":14.0,"y_mm":238.0,"width_mm":90.0,"font_size":10.5,"text_color":"#FFFFFF","align":"left","bold":True,"italic":False,"sort_order":30
    },
    {
        "block_key":"attachments","label":"Anlagen",
        "content_template":"Anlagen:\nLebenslauf / CV\nAnschreiben",
        "x_mm":135.0,"y_mm":272.0,"width_mm":60.0,"font_size":8.5,"text_color":"#FFFFFF","align":"left","bold":False,"italic":False,"sort_order":40
    },
]

def ensure_cover_letter_block_defaults():
    changed=False
    for template in CoverLetterTemplate.query.all():
        existing={b.block_key for b in template.cover_blocks}
        for spec in DEFAULT_COVER_BLOCKS:
            if spec["block_key"] not in existing:
                db.session.add(CoverLetterCoverBlock(template=template,**spec))
                changed=True
    if changed:
        db.session.commit()

def cover_snapshot_values_for_application(application,letter=None):
    base=cover_letter_variables(application)
    profile=CVProfile.query.first()
    cd=application.company_details
    d=application.details
    values=dict(base)
    values.update({
        "applicant_name": (
            letter.sender_name if letter else
            (" ".join([profile.first_name or "",profile.last_name or ""]).strip() if profile else "")
        ) or "",
        "applicant_address": (letter.sender_address if letter else (profile.address if profile else "")) or "",
        "applicant_postal_code": (letter.sender_postal_code if letter else (profile.postal_code if profile else "")) or "",
        "applicant_city": (letter.sender_city if letter else (profile.city if profile else "")) or "",
        "applicant_postal_city": " ".join(x for x in [
            letter.sender_postal_code if letter else (profile.postal_code if profile else ""),
            letter.sender_city if letter else (profile.city if profile else "")
        ] if x),
        "applicant_phone": (letter.sender_phone if letter else (profile.phone if profile else "")) or "",
        "applicant_email": (letter.sender_email if letter else (profile.email if profile else "")) or "",
        "company": (letter.recipient_company if letter else application.company) or "",
        "department": (letter.recipient_department if letter else (cd.department if cd else "")) or "",
        "contact": (letter.recipient_contact if letter else application.contact_person) or "",
        "company_street": (letter.recipient_street if letter else (cd.street if cd else "")) or "",
        "company_postal_code": (letter.recipient_postal_code if letter else (cd.postal_code if cd else "")) or "",
        "company_city": (letter.recipient_city if letter else ((cd.city if cd and cd.city else application.location))) or "",
        "company_postal_city": " ".join(x for x in [
            letter.recipient_postal_code if letter else (cd.postal_code if cd else ""),
            letter.recipient_city if letter else ((cd.city if cd and cd.city else application.location))
        ] if x),
        "position": application.position or "",
        "source": (d.source if d else "") or "",
        "reference": application.reference or "",
        "salary_expectation": application.salary_expectation or "",
        "application_date": application.application_date.strftime("%d.%m.%Y") if application.application_date else "",
        "letter_date": (letter.letter_date.strftime("%d.%m.%Y") if letter and letter.letter_date else date.today().strftime("%d.%m.%Y")),
    })
    return {k:(v or "") for k,v in values.items()}

def create_letter_cover_snapshot(letter):
    values=cover_snapshot_values_for_application(letter.application,letter)
    snap=letter.cover_snapshot
    if not snap:
        snap=ApplicationLetterCoverSnapshot(letter=letter)
        db.session.add(snap)
    snap.data_json=json.dumps(values,ensure_ascii=False)
    return snap

def letter_cover_values(letter):
    if letter.cover_snapshot and letter.cover_snapshot.data_json:
        try:
            data=json.loads(letter.cover_snapshot.data_json)
            if isinstance(data,dict):
                return data
        except Exception:
            pass
    return cover_snapshot_values_for_application(letter.application,letter)

def render_cover_block(block,values):
    return render_letter_placeholders(block.content_template or "",values)

def cover_font_name(block):
    family=block.font_family if block.font_family in ("Helvetica","Times-Roman","Courier") else "Helvetica"
    if family=="Helvetica":
        if block.bold and block.italic: return "Helvetica-BoldOblique"
        if block.bold: return "Helvetica-Bold"
        if block.italic: return "Helvetica-Oblique"
    elif family=="Times-Roman":
        if block.bold and block.italic: return "Times-BoldItalic"
        if block.bold: return "Times-Bold"
        if block.italic: return "Times-Italic"
    elif family=="Courier":
        if block.bold and block.italic: return "Courier-BoldOblique"
        if block.bold: return "Courier-Bold"
        if block.italic: return "Courier-Oblique"
    return family

def wrap_cover_line(text,font_name,font_size,max_width):
    from reportlab.pdfbase.pdfmetrics import stringWidth
    words=(text or "").split()
    if not words: return [""]
    lines=[]; current=words[0]
    for word in words[1:]:
        candidate=current+" "+word
        if stringWidth(candidate,font_name,font_size)<=max_width:
            current=candidate
        else:
            lines.append(current); current=word
    lines.append(current)
    return lines

def draw_cover_blocks(canvas_obj,template,values):
    page_w,page_h=A4
    for block in sorted([b for b in template.cover_blocks if b.enabled],key=lambda b:(b.sort_order,b.id)):
        content=render_cover_block(block,values).strip()
        if not content: continue
        font_name=cover_font_name(block); font_size=max(1.0,min(40.0,float(block.font_size or 12)))
        x=max(0,float(block.x_mm or 0))*mm
        y_top=page_h-max(0,float(block.y_mm or 0))*mm
        max_width=max(15,float(block.width_mm or 60))*mm
        leading=font_size*1.28
        canvas_obj.saveState()
        try:
            canvas_obj.setFillColor(colors.HexColor(block.text_color or "#FFFFFF"))
        except Exception:
            canvas_obj.setFillColor(colors.white)
        canvas_obj.setFont(font_name,font_size)
        y=y_top-font_size
        for raw in content.splitlines():
            wrapped=wrap_cover_line(raw,font_name,font_size,max_width)
            for line in wrapped:
                if y < 4*mm: break
                if block.align=="center":
                    canvas_obj.drawCentredString(x+max_width/2,y,line)
                elif block.align=="right":
                    canvas_obj.drawRightString(x+max_width,y,line)
                else:
                    canvas_obj.drawString(x,y,line)
                y-=leading
        canvas_obj.restoreState()

def save_cover_blocks_from_form(template):
    for block in template.cover_blocks:
        block.enabled=request.form.get(f"block_enabled_{block.id}")=="on"
        block.label=request.form.get(f"block_label_{block.id}",block.label).strip() or block.label
        block.content_template=request.form.get(f"block_content_{block.id}",block.content_template)
        def fval(name,default,lo,hi):
            try: return max(lo,min(hi,float(request.form.get(name,default))))
            except Exception: return default
        block.x_mm=fval(f"block_x_{block.id}",block.x_mm,0,200)
        block.y_mm=fval(f"block_y_{block.id}",block.y_mm,0,292)
        block.width_mm=fval(f"block_width_{block.id}",block.width_mm,15,200)
        block.font_size=fval(f"block_size_{block.id}",block.font_size,1,40)
        family=request.form.get(f"block_font_{block.id}",block.font_family)
        block.font_family=family if family in ("Helvetica","Times-Roman","Courier") else "Helvetica"
        color=request.form.get(f"block_color_{block.id}",block.text_color)
        block.text_color=color if re.match(r"^#[0-9A-Fa-f]{6}$",color or "") else "#FFFFFF"
        align=request.form.get(f"block_align_{block.id}",block.align)
        block.align=align if align in ("left","center","right") else "left"
        block.bold=request.form.get(f"block_bold_{block.id}")=="on"
        block.italic=request.form.get(f"block_italic_{block.id}")=="on"
        block.sort_order=request.form.get(f"block_order_{block.id}",block.sort_order,type=int)

def ensure_cover_letter_cover_defaults():
    changed=False
    for template in CoverLetterTemplate.query.all():
        if not template.cover_style:
            db.session.add(CoverLetterTemplateCover(template=template,enabled=False,fit_mode="cover",background_color="#FFFFFF"))
            changed=True
    if changed: db.session.commit()

def cover_style_for(template):
    style=template.cover_style
    if not style:
        style=CoverLetterTemplateCover(template=template,enabled=False,fit_mode="cover",background_color="#FFFFFF")
        db.session.add(style); db.session.flush()
    return style

def save_cover_style_from_form(style):
    style.enabled=request.form.get("cover_enabled")=="on"
    fit=request.form.get("cover_fit_mode","cover")
    style.fit_mode=fit if fit in ("cover","contain","original_a4") else "cover"
    bg=request.form.get("cover_background_color","#FFFFFF")
    style.background_color=bg if re.match(r"^#[0-9A-Fa-f]{6}$",bg) else "#FFFFFF"

def save_cover_image_upload(style,file_storage):
    if not file_storage or not file_storage.filename: return False
    new_name=save_cv_media_upload(file_storage,f"letter_templates/{style.template_id}","cover")
    if style.image_filename: safe_remove_cv_media(style.image_filename)
    style.image_filename=new_name
    return True

def draw_cover_image_page(canvas_obj,cover):
    from reportlab.lib.utils import ImageReader
    page_w,page_h=A4
    canvas_obj.saveState()
    canvas_obj.setFillColor(colors.HexColor(cover.background_color or "#FFFFFF"))
    canvas_obj.rect(0,0,page_w,page_h,stroke=0,fill=1)
    path=cv_media_path(cover.image_filename)
    if path:
        reader=ImageReader(str(path)); iw,ih=reader.getSize()
        if iw and ih:
            if cover.fit_mode=="contain":
                scale=min(page_w/float(iw),page_h/float(ih)); w=iw*scale; h=ih*scale
                canvas_obj.drawImage(reader,(page_w-w)/2,(page_h-h)/2,width=w,height=h,preserveAspectRatio=True,mask="auto")
            elif cover.fit_mode=="original_a4":
                canvas_obj.drawImage(reader,0,0,width=page_w,height=page_h,preserveAspectRatio=False,mask="auto")
            else:
                scale=max(page_w/float(iw),page_h/float(ih)); w=iw*scale; h=ih*scale
                canvas_obj.saveState(); path_clip=canvas_obj.beginPath(); path_clip.rect(0,0,page_w,page_h); canvas_obj.clipPath(path_clip,stroke=0,fill=0)
                canvas_obj.drawImage(reader,(page_w-w)/2,(page_h-h)/2,width=w,height=h,preserveAspectRatio=True,mask="auto")
                canvas_obj.restoreState()
    canvas_obj.restoreState()

def build_cover_page_pdf(cover,letter):
    if not cover.enabled or not cover.image_filename or not cv_media_path(cover.image_filename): return None
    from reportlab.pdfgen import canvas as pdfcanvas
    buf=io.BytesIO(); c=pdfcanvas.Canvas(buf,pagesize=A4)
    draw_cover_image_page(c,cover)
    draw_cover_blocks(c,cover.template,letter_cover_values(letter))
    c.showPage(); c.save(); buf.seek(0); return buf

def prepend_cover_to_pdf(body_buffer,cover,letter):
    cover_buffer=build_cover_page_pdf(cover,letter)
    if not cover_buffer:
        body_buffer.seek(0); return body_buffer
    writer=PdfWriter()
    for page in PdfReader(cover_buffer).pages: writer.add_page(page)
    body_buffer.seek(0)
    for page in PdfReader(body_buffer).pages: writer.add_page(page)
    merged=io.BytesIO(); writer.write(merged); merged.seek(0); return merged


def ensure_cover_letter_footer_defaults():
    changed=False
    for template in CoverLetterTemplate.query.all():
        if not template.footer_style:
            db.session.add(CoverLetterFooterStyle(
                template=template,
                enabled=(template.slug=="modern"),
                layout="centered",
                show_divider=True,
                divider_color="#D0D5DD",
                divider_width_pct=90,
                logo_height_mm=14.0,
                logo_gap_mm=16.0,
                primary_color=template.accent_color or "#5B6E6A",
                secondary_color="#E5E7EB",
            ))
            changed=True
    if changed: db.session.commit()

def footer_style_for(template):
    style=template.footer_style
    if not style:
        style=CoverLetterFooterStyle(template=template,enabled=(template.slug=="modern"))
        db.session.add(style); db.session.flush()
    return style

def template_certification_links(template, enabled_only=False):
    q=CoverLetterTemplateCertification.query.filter_by(template_id=template.id)
    if enabled_only: q=q.filter_by(enabled=True)
    return q.order_by(CoverLetterTemplateCertification.sort_order.asc(),CoverLetterTemplateCertification.id.asc()).all()

def template_certifications(template):
    return [link.certification for link in template_certification_links(template,True) if link.certification and link.certification.is_active]

def save_footer_style_from_form(style):
    style.enabled=request.form.get("footer_enabled")=="on"
    style.layout="centered"
    style.show_divider=request.form.get("footer_show_divider")=="on"
    divider=request.form.get("footer_divider_color","#D0D5DD")
    style.divider_color=divider if re.match(r"^#[0-9A-Fa-f]{6}$",divider) else "#D0D5DD"
    primary=request.form.get("footer_primary_color","#5B6E6A")
    secondary=request.form.get("footer_secondary_color","#E5E7EB")
    style.primary_color=primary if re.match(r"^#[0-9A-Fa-f]{6}$",primary) else "#5B6E6A"
    style.secondary_color=secondary if re.match(r"^#[0-9A-Fa-f]{6}$",secondary) else "#E5E7EB"
    style.divider_width_pct=max(50,min(100,request.form.get("footer_divider_width_pct",90,type=int)))
    try: style.logo_height_mm=max(7.0,min(22.0,float(request.form.get("footer_logo_height_mm","14"))))
    except (TypeError,ValueError): style.logo_height_mm=14.0
    try: style.logo_gap_mm=max(3.0,min(28.0,float(request.form.get("footer_logo_gap_mm","16"))))
    except (TypeError,ValueError): style.logo_gap_mm=16.0

def save_template_certification_selection(template):
    existing={link.certification_id:link for link in template.footer_certification_links}
    for cert in CertificationLogo.query.order_by(CertificationLogo.sort_order,CertificationLogo.id).all():
        link=existing.get(cert.id)
        selected=request.form.get(f"cert_enabled_{cert.id}")=="on"
        order=request.form.get(f"cert_order_{cert.id}",cert.sort_order,type=int)
        if not link:
            link=CoverLetterTemplateCertification(template=template,certification=cert,enabled=selected,sort_order=order)
            db.session.add(link)
        else:
            link.enabled=selected; link.sort_order=order

def certification_image_dimensions(cert, target_height_pt):
    path=cv_media_path(cert.filename)
    if not path: return None
    try:
        from reportlab.lib.utils import ImageReader
        reader=ImageReader(str(path)); w,h=reader.getSize()
        if not w or not h: return None
        width=target_height_pt*(float(w)/float(h))
        max_width=42*mm
        if width>max_width:
            ratio=max_width/width; width=max_width; target_height_pt*=ratio
        return path,width,target_height_pt
    except Exception:
        current_app.logger.exception("Zertifizierungslogo konnte nicht gelesen werden: %s",path)
        return None

def draw_cover_letter_footer(canvas, doc, letter):
    style=footer_style_for(letter.template)
    certs=template_certifications(letter.template)
    if not style.enabled or not certs: return
    page_w,page_h=A4
    usable=page_w-doc.leftMargin-doc.rightMargin
    center_x=page_w/2
    line_y=26*mm
    if style.show_divider:
        line_w=usable*(style.divider_width_pct/100.0)
        canvas.saveState(); canvas.setStrokeColor(colors.HexColor(style.divider_color or "#D0D5DD")); canvas.setLineWidth(0.6); canvas.line(center_x-line_w/2,line_y,center_x+line_w/2,line_y); canvas.restoreState()
    target_h=style.logo_height_mm*mm
    items=[]
    for cert in certs:
        dims=certification_image_dimensions(cert,target_h)
        if dims: items.append((cert,)+dims)
    if not items: return
    gap=style.logo_gap_mm*mm
    total=sum(x[2] for x in items)+gap*max(0,len(items)-1)
    max_total=usable*0.94
    if total>max_total:
        factor=max_total/total
        items=[(cert,path,w*factor,h*factor) for cert,path,w,h in items]
        gap*=factor; total=sum(x[2] for x in items)+gap*max(0,len(items)-1)
    x=center_x-total/2
    base_y=7*mm
    canvas.saveState()
    for cert,path,w,h in items:
        try: canvas.drawImage(str(path),x,base_y,width=w,height=h,preserveAspectRatio=True,mask='auto',anchor='c')
        except Exception: current_app.logger.exception("Zertifizierungslogo konnte nicht ins PDF gezeichnet werden: %s",path)
        x+=w+gap
    canvas.restoreState()

def letter_render_metrics(template):
    """Single source of truth for cover-letter preview/PDF proportions."""
    scale=max(0.65,min(1.6,float(template.font_scale or 1.0)))
    compact=template.slug=="compact"
    return {
        "scale": scale,
        "sender_pt": 8.0*scale,
        "sender_leading_pt": 10.0*scale,
        "address_pt": 9.0*scale,
        "address_leading_pt": 11.0*scale,
        "subject_pt": 11.0*scale,
        "subject_leading_pt": 14.0*scale,
        "body_pt": (9.0 if compact else 10.0)*scale,
        "body_leading_pt": (12.0 if compact else 14.0)*scale,
        "signature_width_mm": 42.0,
        "signature_height_mm": 20.0,
        "sender_gap_mm": 7.0,
        "recipient_gap_mm": 9.0,
        "date_gap_mm": 7.0,
        "paragraph_gap_mm": 4.0,
        "signature_gap_mm": 1.0,
    }

def build_application_letter_pdf(letter):
    buffer=io.BytesIO(); t=letter.template
    metrics=letter_render_metrics(t)
    margin=max(12,min(30,t.page_margin_mm))*mm
    footer=footer_style_for(t); footer_active=footer.enabled and bool(template_certifications(t))
    bottom_margin=(38*mm if footer_active else 16*mm)
    doc=SimpleDocTemplate(buffer,pagesize=A4,rightMargin=margin,leftMargin=margin,topMargin=16*mm,bottomMargin=bottom_margin,title=letter.subject or letter.title,author=letter.sender_name or get_app_settings().get("author",""))
    styles=getSampleStyleSheet(); accent=colors.HexColor(t.accent_color or "#172033"); font=t.font_family if t.font_family in ("Helvetica","Times-Roman","Courier") else "Helvetica"
    styles.add(ParagraphStyle(name="LetterSender",parent=styles["Normal"],fontName=font,fontSize=metrics["sender_pt"],leading=metrics["sender_leading_pt"],textColor=colors.HexColor("#667085")))
    styles.add(ParagraphStyle(name="LetterAddress",parent=styles["Normal"],fontName=font,fontSize=metrics["address_pt"],leading=metrics["address_leading_pt"]))
    styles.add(ParagraphStyle(name="LetterSubject",parent=styles["Heading2"],fontName=font,fontSize=metrics["subject_pt"],leading=metrics["subject_leading_pt"],textColor=accent,spaceAfter=metrics["date_gap_mm"]*mm))
    styles.add(ParagraphStyle(name="LetterBody",parent=styles["Normal"],fontName=font,fontSize=metrics["body_pt"],leading=metrics["body_leading_pt"],spaceAfter=metrics["paragraph_gap_mm"]*mm))
    story=[]
    sender_line=letter_sender_line(letter)
    story += [Paragraph(escape(sender_line),styles["LetterSender"]),Spacer(1,metrics["sender_gap_mm"]*mm)]
    recipient=[letter.recipient_company,letter.recipient_department,(f"z. Hd. {letter.recipient_contact}" if letter.recipient_contact else None),letter.recipient_street," ".join(x for x in [letter.recipient_postal_code,letter.recipient_city] if x),letter.recipient_country]
    story.append(Paragraph("<br/>".join(escape(x) for x in recipient if x),styles["LetterAddress"])); story.append(Spacer(1,metrics["recipient_gap_mm"]*mm))
    date_line=letter.letter_date.strftime("%d.%m.%Y") if letter.letter_date else ""
    if letter.sender_city: date_line=f"{letter.sender_city}, {date_line}"
    story.append(Paragraph(escape(date_line),ParagraphStyle(name="LetterDate",parent=styles["LetterAddress"],alignment=2))); story.append(Spacer(1,metrics["date_gap_mm"]*mm))
    story.append(Paragraph(escape(letter.subject or "Anschreiben"),styles["LetterSubject"])); story.append(Paragraph(escape(letter.salutation or "Sehr geehrte Damen und Herren,"),styles["LetterBody"]))
    for para in letter_paragraphs(letter): story.append(Paragraph(pdf_text(para),styles["LetterBody"]))
    story.append(Spacer(1,2*mm)); story.append(Paragraph(escape(letter.signoff or "Mit freundlichen Grüßen"),styles["LetterBody"])); story.append(Spacer(1,2*mm))
    sig=cv_media_path(letter.signature_filename)
    img=reportlab_image(sig,metrics["signature_width_mm"],metrics["signature_height_mm"]) if sig else None
    if img:
        # RLImage defaults may center depending on surrounding flowables/styles.
        # The HTML preview is explicitly left-aligned, so make PDF alignment explicit.
        img.hAlign="LEFT"
        story.append(img)
        story.append(Spacer(1,metrics["signature_gap_mm"]*mm))
    if letter.sender_name: story.append(Paragraph(escape(letter.sender_name),styles["LetterBody"]))
    if footer_active:
        doc.build(story,onFirstPage=lambda c,d: draw_cover_letter_footer(c,d,letter),onLaterPages=lambda c,d: draw_cover_letter_footer(c,d,letter))
    else:
        doc.build(story)
    buffer.seek(0)
    return prepend_cover_to_pdf(buffer,cover_style_for(t),letter)

APP_VERSION = "v8.5.5"

DEFAULT_MENU_ITEMS = [
    {"key":"dashboard","label":"Dashboard","endpoint":"dashboard","group":"Bewerbungen","order":10,"visible":True},
    {"key":"overview","label":"Übersicht","endpoint":"index","group":"Bewerbungen","order":20,"visible":True},
    {"key":"new_application","label":"+ Bewerbung","endpoint":"new_application","group":"Bewerbungen","order":30,"visible":True},
    {"key":"calendar","label":"Kalender","endpoint":"calendar_view","group":"Bewerbungen","order":40,"visible":True},
    {"key":"reminders","label":"Erinnerungen","endpoint":"reminders","group":"Bewerbungen","order":50,"visible":True},
    {"key":"kanban","label":"Kanban","endpoint":"kanban","group":"Bewerbungen","order":60,"visible":True},
    {"key":"cv","label":"Lebenslauf","endpoint":"cv_index","group":"Lebenslauf","order":10,"visible":True},
    {"key":"cv_templates","label":"CV-Templates","endpoint":"cv_templates","group":"Lebenslauf","order":20,"visible":True},
    {"key":"universal_cvs","label":"Universelle CVs","endpoint":"universal_cvs","group":"Lebenslauf","order":25,"visible":True},
    {"key":"cover_letter","label":"Anschreiben","endpoint":"cover_letter_master","group":"Lebenslauf","order":30,"visible":True},
    {"key":"certification_logos","label":"Zertifizierungslogos","endpoint":"certification_logos","group":"Lebenslauf","order":40,"visible":True},
    {"key":"export","label":"Excel","endpoint":"export_excel","group":"Werkzeuge","order":10,"visible":True},
    {"key":"backup","label":"Backup","endpoint":"backups","group":"Werkzeuge","order":20,"visible":True},
    {"key":"statuses","label":"Status","endpoint":"statuses_page","group":"System","order":10,"visible":True},
    {"key":"personalize","label":"Personalisieren","endpoint":"personalize","group":"System","order":20,"visible":True},
]

def get_menu_config():
    setting=AppSetting.query.filter_by(key="menu_config").first()
    if not setting or not setting.value:
        return [dict(item) for item in DEFAULT_MENU_ITEMS]
    try:
        saved=json.loads(setting.value)
    except (TypeError,ValueError):
        saved=[]
    by_key={x.get("key"):x for x in saved if isinstance(x,dict)}
    result=[]
    for default in DEFAULT_MENU_ITEMS:
        item=dict(default)
        item.update({k:v for k,v in by_key.get(default["key"],{}).items() if k in {"label","group","order","visible"}})
        result.append(item)
    return result

def save_menu_config_from_form():
    items=get_menu_config()
    updated=[]
    for item in items:
        key=item["key"]
        label=request.form.get(f"menu_label_{key}",item["label"]).strip() or item["label"]
        group=request.form.get(f"menu_group_{key}",item["group"]).strip() or "Sonstiges"
        order=request.form.get(f"menu_order_{key}",item["order"],type=int)
        visible=request.form.get(f"menu_visible_{key}")=="on"
        updated.append({**item,"label":label,"group":group,"order":order,"visible":visible})
    set_app_setting("menu_config",json.dumps(updated,ensure_ascii=False))

def grouped_menu():
    groups={}
    group_order={"Bewerbungen":10,"Lebenslauf":20,"Werkzeuge":30,"System":40}
    for item in get_menu_config():
        if not item.get("visible",True):
            continue
        groups.setdefault(item["group"],[]).append(item)
    result=[]
    for group,items in groups.items():
        items=sorted(items,key=lambda x:(int(x.get("order",0)),x.get("label","")))
        result.append({"name":group,"entries":items,"order":group_order.get(group,100)})
    return sorted(result,key=lambda x:(x["order"],x["name"]))

def ensure_cv_template_styles():
    changed=False
    defaults={"classic":"classic","modern":"modern","compact":"compact"}
    for template in CVTemplate.query.all():
        if not template.style:
            base=defaults.get(template.slug,"classic")
            accent={"classic":"#172033","modern":"#2f5d8a","compact":"#3d4653"}.get(base,"#172033")
            db.session.add(CVTemplateStyle(template=template,base_layout=base,accent_color=accent))
            changed=True
    if changed:
        db.session.commit()

def template_style(template):
    if template and template.style:
        return template.style
    class Fallback:
        base_layout="classic"; accent_color="#172033"; font_family="Helvetica"; font_scale=1.0
        page_margin_mm=15; show_logo=True; show_company=True; show_target_position=True
        section_order="experience,education,projects,skills,languages,certifications,other"
    return Fallback()

def cv_render_metrics(style):
    """Single source of truth for CV preview/PDF proportions."""
    scale=max(0.65,min(1.6,float(style.font_scale or 1.0)))
    compact=style.base_layout=="compact"
    return {
        "scale": scale,
        "body_pt": (8.5 if compact else 9.5)*scale,
        "line_height": 1.4 if not compact else 1.28,
        "name_pt": (18 if compact else 22)*scale,
        "target_pt": 11*scale,
        "contact_pt": 8.5*scale,
        "document_pt": 8*scale,
        "section_pt": 11*scale,
        "entry_title_pt": 10*scale,
        "small_pt": 7.5*scale,
        "modern_sidebar_mm": 48.0,
        "modern_gap_mm": 8.0,
        "modern_header_mm": 58.0,
        "header_photo_w_mm": 30.0,
        "header_photo_h_mm": 38.0,
        "header_logo_mm": 19.0,
    }

def normalized_section_order(style):
    allowed=list(CV_SECTION_LABELS.keys())
    raw=[x.strip() for x in (style.section_order or "").split(",") if x.strip() in allowed]
    for key in allowed:
        if key not in raw:
            raw.append(key)
    return raw

DEFAULT_CV_TEMPLATES = [
    ("Klassisch","classic","Konservatives, einspaltiges Layout mit klarer Typografie.",10),
    ("Modern","modern","Moderner Kopfbereich und zweispaltige Darstellung für Skills und Sprachen.",20),
    ("Kompakt","compact","Platzsparendes Layout für kurze, fokussierte Lebensläufe.",30),
]

CV_SECTION_LABELS = {
    "experience":"Berufserfahrung",
    "education":"Ausbildung & Studium",
    "skills":"Kenntnisse & Skills",
    "languages":"Sprachen",
    "certifications":"Zertifikate & Weiterbildungen",
    "projects":"Projekte",
    "other":"Weitere Stationen",
}

def ensure_default_cv_templates():
    changed=False
    for name,slug,description,sort_order in DEFAULT_CV_TEMPLATES:
        item=CVTemplate.query.filter_by(slug=slug).first()
        if not item:
            db.session.add(CVTemplate(name=name,slug=slug,description=description,sort_order=sort_order,is_active=True))
            changed=True
    if changed:
        db.session.commit()

def serialize_cv_value(value):
    if isinstance(value,(date,datetime)):
        return value.isoformat()
    return value

def snapshot_model_item(item, fields):
    data={}
    for name, _label, _field_type, _required in fields:
        data[name]=serialize_cv_value(getattr(item,name,None))
    return data

def snapshot_profile(profile):
    if not profile:
        return {}
    keys=[
        "first_name","last_name","professional_title","birth_date","birth_place",
        "address","postal_code","city","country","email","phone","website",
        "linkedin","xing","github","summary"
    ]
    return {key:serialize_cv_value(getattr(profile,key,None)) for key in keys}

def application_cv_profile(application_cv):
    try:
        return json.loads(application_cv.profile_json or "{}")
    except (TypeError,ValueError):
        return {}

def application_cv_entry_data(entry):
    try:
        return json.loads(entry.data_json or "{}")
    except (TypeError,ValueError):
        return {}

def cv_entry_groups(application_cv, visible_only=False):
    groups={key:[] for key in CV_SECTION_LABELS}
    for entry in sorted(application_cv.entries,key=lambda e:(e.section,e.sort_order,e.id)):
        if visible_only and not entry.visible:
            continue
        if entry.section in groups:
            groups[entry.section].append(entry)
    return groups

def create_application_cv_snapshot(application, template, title=None, target_company=None, target_position=None, profile_summary=None):
    profile=CVProfile.query.first()
    cv=ApplicationCV(
        application=application,
        template=template,
        title=(title or "Lebenslauf").strip() or "Lebenslauf",
        target_company=(target_company or application.company or "").strip() or None,
        target_position=(target_position or application.position or "").strip() or None,
        profile_summary=(profile_summary if profile_summary is not None else (profile.summary if profile else None)),
        profile_json=json.dumps(snapshot_profile(profile),ensure_ascii=False),
    )
    db.session.add(cv)
    db.session.flush()
    ensure_application_cv_header(cv)
    for section,cfg in CV_SECTIONS.items():
        model=cfg["model"]
        if hasattr(model,"start_date"):
            items=model.query.order_by(model.sort_order.asc(),model.start_date.desc(),model.id.desc()).all()
        else:
            items=model.query.order_by(model.sort_order.asc(),model.id.asc()).all()
        for index,item in enumerate(items):
            db.session.add(ApplicationCVEntry(
                application_cv=cv,
                section=section,
                source_id=item.id,
                sort_order=(getattr(item,"sort_order",None) if getattr(item,"sort_order",None) is not None else index),
                visible=True,
                data_json=json.dumps(snapshot_model_item(item,cfg["fields"]),ensure_ascii=False),
            ))
    return cv

def create_universal_cv_snapshot(template,title=None,target_position=None,profile_summary=None):
    profile=CVProfile.query.first()
    cv=UniversalCV(
        template=template,
        title=(title or "Allgemeiner Lebenslauf").strip() or "Allgemeiner Lebenslauf",
        target_company=None,
        target_position=(target_position or "").strip() or None,
        profile_summary=(profile_summary if profile_summary is not None else (profile.summary if profile else None)),
        profile_json=json.dumps(snapshot_profile(profile),ensure_ascii=False),
    )
    db.session.add(cv); db.session.flush()
    ensure_application_cv_header(cv)
    for section,cfg in CV_SECTIONS.items():
        model=cfg["model"]
        if hasattr(model,"start_date"):
            items=model.query.order_by(model.sort_order.asc(),model.start_date.desc(),model.id.desc()).all()
        else:
            items=model.query.order_by(model.sort_order.asc(),model.id.asc()).all()
        for index,item in enumerate(items):
            db.session.add(UniversalCVEntry(
                universal_cv=cv,section=section,source_id=item.id,
                sort_order=(getattr(item,"sort_order",None) if getattr(item,"sort_order",None) is not None else index),
                visible=True,data_json=json.dumps(snapshot_model_item(item,cfg["fields"]),ensure_ascii=False),
            ))
    return cv

def copy_universal_cv_to_application(universal,application,template=None):
    cv=ApplicationCV(
        application=application,
        template=template or universal.template,
        title=universal.title or "Lebenslauf",
        target_company=application.company,
        target_position=application.position or universal.target_position,
        profile_summary=universal.profile_summary,
        profile_json=universal.profile_json,
    )
    db.session.add(cv); db.session.flush()
    source_header=ensure_application_cv_header(universal)
    header=ensure_application_cv_header(cv)
    for attr in ("document_title","header_subtitle","header_layout","show_document_title","show_professional_title","show_target_position","show_contact"):
        setattr(header,attr,getattr(source_header,attr))
    header.show_target_company=True
    if source_header.logo_filename:
        header.logo_filename=copy_cv_media(source_header.logo_filename,f"snapshots/{cv.id}","logo")
    if source_header.photo_filename:
        header.photo_filename=copy_cv_media(source_header.photo_filename,f"snapshots/{cv.id}","photo")
    for entry in universal.entries:
        db.session.add(ApplicationCVEntry(
            application_cv=cv,section=entry.section,source_id=entry.source_id,
            sort_order=entry.sort_order,visible=entry.visible,data_json=entry.data_json,
        ))
    return cv

def cv_media_root():
    root=Path(current_app.config["UPLOAD_DIR"]) / "cv_media"
    root.mkdir(parents=True,exist_ok=True)
    return root

def master_cv_photo_filename():
    setting=AppSetting.query.filter_by(key="cv_profile_photo_filename").first()
    return (setting.value or "") if setting else ""

def cv_media_path(relative_name):
    if not relative_name:
        return None
    path=cv_media_root() / relative_name
    return path if path.exists() else None

def safe_remove_cv_media(relative_name):
    path=cv_media_path(relative_name)
    if path and path.is_file():
        path.unlink()

def save_cv_media_upload(file_storage, subdir, prefix):
    if not file_storage or not file_storage.filename:
        return None
    ext=file_storage.filename.rsplit(".",1)[1].lower() if "." in file_storage.filename else ""
    if ext not in LOGO_EXTENSIONS:
        raise ValueError("Erlaubte Bildformate: PNG, JPG/JPEG und WEBP.")
    target_dir=cv_media_root() / subdir
    target_dir.mkdir(parents=True,exist_ok=True)
    filename=f"{prefix}_{uuid.uuid4().hex}.{ext}"
    file_storage.save(target_dir / filename)
    return f"{subdir}/{filename}"

def copy_cv_media(source_relative, subdir, prefix):
    source=cv_media_path(source_relative)
    if not source:
        return None
    target_dir=cv_media_root() / subdir
    target_dir.mkdir(parents=True,exist_ok=True)
    ext=source.suffix.lower() or ".png"
    filename=f"{prefix}_{uuid.uuid4().hex}{ext}"
    shutil.copy2(source,target_dir / filename)
    return f"{subdir}/{filename}"

def system_logo_relative():
    settings=get_app_settings()
    filename=settings.get("logo_filename","")
    if not filename:
        return None
    source=Path(current_app.config["UPLOAD_DIR"]) / "branding" / filename
    if not source.exists():
        return None
    target_dir=cv_media_root() / "system"
    target_dir.mkdir(parents=True,exist_ok=True)
    ext=source.suffix.lower() or ".png"
    target=target_dir / f"systemlogo_{uuid.uuid4().hex}{ext}"
    shutil.copy2(source,target)
    return f"system/{target.name}"

def ensure_application_cv_header(application_cv):
    header=application_cv.header
    if header:
        return header
    common=dict(
        document_title="Lebenslauf",
        header_subtitle=None,
        header_layout="photo_right",
        show_document_title=True,
        show_professional_title=True,
        show_target_position=True,
        show_contact=True,
    )
    if isinstance(application_cv,UniversalCV):
        header=UniversalCVHeader(universal_cv=application_cv,show_target_company=False,**common)
    else:
        header=ApplicationCVHeader(application_cv=application_cv,show_target_company=True,**common)
    db.session.add(header)
    db.session.flush()
    return header

def configure_new_cv_header(cv):
    header=ensure_application_cv_header(cv)
    header.document_title=request.form.get("document_title","Lebenslauf").strip() or "Lebenslauf"
    header.header_subtitle=request.form.get("header_subtitle","").strip() or None
    layout=request.form.get("header_layout","photo_right")
    header.header_layout=layout if layout in ("photo_right","photo_left","centered","minimal") else "photo_right"
    header.show_document_title=request.form.get("show_document_title")=="on"
    header.show_professional_title=request.form.get("show_professional_title")=="on"
    header.show_target_company=request.form.get("show_target_company")=="on"
    header.show_target_position=request.form.get("show_target_position")=="on"
    header.show_contact=request.form.get("show_contact")=="on"

    subdir=(f"universal/{cv.id}" if isinstance(cv,UniversalCV) else f"snapshots/{cv.id}")

    logo_source=request.form.get("logo_source","none")
    if logo_source=="system":
        header.logo_filename=system_logo_relative()
    elif logo_source=="custom":
        header.logo_filename=save_cv_media_upload(request.files.get("header_logo"),subdir,"logo")
    else:
        header.logo_filename=None

    photo_source=request.form.get("photo_source","none")
    if photo_source=="master":
        header.photo_filename=copy_cv_media(master_cv_photo_filename(),subdir,"photo")
    elif photo_source=="custom":
        header.photo_filename=save_cv_media_upload(request.files.get("header_photo"),subdir,"photo")
    else:
        header.photo_filename=None
    return header

def cv_period(data):
    start=data.get("start_date")
    end=data.get("end_date")
    current=data.get("is_current")
    def fmt(v):
        if not v:
            return ""
        try:
            return datetime.fromisoformat(v).strftime("%m.%Y")
        except Exception:
            return str(v)
    left=fmt(start)
    right="heute" if current else fmt(end)
    if left and right:
        return f"{left} – {right}"
    return left or right or ""

def reportlab_image(path_value, max_width_mm, max_height_mm):
    """Create a ReportLab image safely. Invalid/unsupported images are skipped."""
    if not path_value:
        return None
    try:
        from reportlab.lib.utils import ImageReader
        reader=ImageReader(str(path_value))
        width,height=reader.getSize()
        if not width or not height:
            return None
        max_w=max_width_mm*mm
        max_h=max_height_mm*mm
        ratio=min(max_w/float(width),max_h/float(height))
        draw_w=max(1,width*ratio)
        draw_h=max(1,height*ratio)
        return RLImage(str(path_value),width=draw_w,height=draw_h)
    except Exception:
        current_app.logger.exception("CV-PDF: Bild konnte nicht geladen werden: %s",path_value)
        return None

def cv_vertical_stack(items,width,align="LEFT"):
    """Render CV header flowables as explicit vertical rows."""
    visible=[item for item in items if item is not None]
    if not visible:
        return Spacer(1,0)
    stack=Table([[item] for item in visible],colWidths=[width])
    stack.setStyle(TableStyle([
        ("VALIGN",(0,0),(-1,-1),"TOP"),
        ("ALIGN",(0,0),(-1,-1),align),
        ("LEFTPADDING",(0,0),(-1,-1),0),
        ("RIGHTPADDING",(0,0),(-1,-1),0),
        ("TOPPADDING",(0,0),(-1,-1),0),
        ("BOTTOMPADDING",(0,0),(-1,-1),0),
    ]))
    return stack

def build_application_cv_pdf(application_cv):
    settings=get_app_settings()
    profile=application_cv_profile(application_cv)
    header=ensure_application_cv_header(application_cv)
    groups=cv_entry_groups(application_cv,visible_only=True)
    style=template_style(application_cv.template)
    metrics=cv_render_metrics(style)
    template_slug=style.base_layout
    section_order=normalized_section_order(style)
    margin=style.page_margin_mm*mm

    buffer=io.BytesIO()
    page_width,page_height=A4
    usable_width=page_width-(2*margin)
    sidebar_width=min(metrics["modern_sidebar_mm"]*mm,usable_width*0.32) if template_slug=="modern" else 0
    sidebar_gap=metrics["modern_gap_mm"]*mm if template_slug=="modern" else 0
    main_width=usable_width-sidebar_width-sidebar_gap if template_slug=="modern" else usable_width
    if template_slug=="modern":
        doc=BaseDocTemplate(
            buffer,pagesize=A4,rightMargin=margin,leftMargin=margin,
            topMargin=0,bottomMargin=margin,
            title=f"{application_cv.title} - {application_cv.target_position or ''}",
            author=settings.get("author","Peter Lange"),
        )
    else:
        doc=SimpleDocTemplate(
            buffer,pagesize=A4,rightMargin=margin,leftMargin=margin,
            topMargin=margin,bottomMargin=margin,
            title=f"{application_cv.title} - {application_cv.target_position or ''}",
            author=settings.get("author","Peter Lange"),
        )
    layout_width=main_width if template_slug=="modern" else usable_width
    styles=getSampleStyleSheet()
    accent=colors.HexColor(style.accent_color)
    scale=metrics["scale"]
    font=style.font_family
    styles.add(ParagraphStyle(name="CVName",parent=styles["Title"],fontName=font,fontSize=metrics["name_pt"],leading=metrics["name_pt"]*1.09,textColor=colors.HexColor("#202733"),spaceAfter=2,alignment=TA_LEFT))
    styles.add(ParagraphStyle(name="CVTarget",parent=styles["Heading2"],fontName=font,fontSize=metrics["target_pt"],leading=metrics["target_pt"]*1.18,textColor=colors.HexColor("#526172"),spaceAfter=2.8,alignment=TA_LEFT))
    styles.add(ParagraphStyle(name="CVSection",parent=styles["Heading2"],fontName=font,fontSize=metrics["section_pt"],leading=metrics["section_pt"]*1.18,textColor=accent,spaceBefore=5*mm,spaceAfter=3*mm))
    styles.add(ParagraphStyle(name="CVItemTitle",parent=styles["Normal"],fontName=font,fontSize=metrics["entry_title_pt"],leading=metrics["entry_title_pt"]*1.18,spaceAfter=1))
    styles.add(ParagraphStyle(name="CVBody",parent=styles["Normal"],fontName=font,fontSize=metrics["body_pt"],leading=metrics["body_pt"]*metrics["line_height"]))
    styles.add(ParagraphStyle(name="CVSmall",parent=styles["Normal"],fontName=font,fontSize=metrics["small_pt"],leading=metrics["small_pt"]*1.2,textColor=colors.HexColor("#667085"),alignment=TA_LEFT))
    styles.add(ParagraphStyle(name="CVDocument",parent=styles["Normal"],fontName=font,fontSize=metrics["document_pt"],leading=metrics["document_pt"]*1.15,textColor=colors.HexColor("#667085"),spaceAfter=2,alignment=TA_LEFT))
    styles.add(ParagraphStyle(name="CVNameCenter",parent=styles["CVName"],alignment=TA_CENTER))
    styles.add(ParagraphStyle(name="CVTargetCenter",parent=styles["CVTarget"],alignment=TA_CENTER))
    styles.add(ParagraphStyle(name="CVSmallCenter",parent=styles["CVSmall"],alignment=TA_CENTER))
    styles.add(ParagraphStyle(name="CVDocumentCenter",parent=styles["CVDocument"],alignment=TA_CENTER))
    styles.add(ParagraphStyle(name="CVFooter",parent=styles["Normal"],fontName=font,fontSize=metrics["small_pt"],leading=metrics["small_pt"]*1.2,textColor=colors.HexColor("#77808e"),alignment=TA_CENTER))
    story=[]

    logo_path_value=cv_media_path(header.logo_filename)
    photo_path_value=cv_media_path(header.photo_filename)
    logo=reportlab_image(logo_path_value,metrics["header_logo_mm"],metrics["header_logo_mm"])
    photo=reportlab_image(photo_path_value,metrics["header_photo_w_mm"],metrics["header_photo_h_mm"])

    full_name=(" ".join([profile.get("first_name") or "",profile.get("last_name") or ""])).strip() or "Lebenslauf"
    target_parts=[]
    if header.show_target_position and application_cv.target_position: target_parts.append(application_cv.target_position)
    if header.show_target_company and application_cv.target_company: target_parts.append(application_cv.target_company)
    target_line=" · ".join(target_parts)

    centered_header=header.header_layout=="centered"
    document_style=styles["CVDocumentCenter"] if centered_header else styles["CVDocument"]
    name_style=styles["CVNameCenter"] if centered_header else styles["CVName"]
    target_style=styles["CVTargetCenter"] if centered_header else styles["CVTarget"]
    small_style=styles["CVSmallCenter"] if centered_header else styles["CVSmall"]

    header_text=[]
    if header.show_document_title:
        header_text.append(Paragraph(escape((header.document_title or "Lebenslauf").upper()),document_style))
    header_text.append(Paragraph(escape(full_name),name_style))
    if header.show_professional_title and profile.get("professional_title"):
        header_text.append(Paragraph(escape(profile.get("professional_title")),target_style))
    if header.header_subtitle:
        header_text.append(Paragraph(escape(header.header_subtitle),target_style))
    if target_line:
        header_text.append(Paragraph(escape(target_line),target_style))
    if header.show_contact:
        contact=[x for x in [profile.get("email"),profile.get("phone")," ".join([profile.get("postal_code") or "",profile.get("city") or ""]).strip()] if x]
        if contact:
            header_text.append(Paragraph(escape(" · ".join(contact)),small_style))

    if template_slug=="modern":
        # Full-width first-page header. Each textual element is a dedicated row,
        # so ReportLab mirrors the vertical HTML header structure.
        header_bg=colors.HexColor("#eef4f9")
        photo_col=38*mm
        if header.header_layout=="centered":
            center_items=[]
            if logo: center_items.append(logo)
            center_items.extend(header_text)
            if photo: center_items.append(photo)
            text_stack=cv_vertical_stack(center_items,page_width-(2*margin),"CENTER")
            modern_header=Table([[text_stack]],colWidths=[page_width])
            modern_header.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,-1),header_bg),("ALIGN",(0,0),(-1,-1),"CENTER"),
                ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                ("LEFTPADDING",(0,0),(-1,-1),margin),("RIGHTPADDING",(0,0),(-1,-1),margin),
                ("TOPPADDING",(0,0),(-1,-1),10*mm),("BOTTOMPADDING",(0,0),(-1,-1),6*mm),
            ]))
        elif header.header_layout=="photo_left" and photo:
            content_items=[]
            if logo: content_items.append(logo)
            content_items.extend(header_text)
            content_stack=cv_vertical_stack(content_items,page_width-photo_col-margin-5*mm,"LEFT")
            modern_header=Table([[photo,content_stack]],colWidths=[photo_col,page_width-photo_col])
            modern_header.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,-1),header_bg),("VALIGN",(0,0),(-1,-1),"TOP"),
                ("LEFTPADDING",(0,0),(0,0),margin),("RIGHTPADDING",(1,0),(1,0),margin),
                ("LEFTPADDING",(1,0),(1,0),5*mm),("RIGHTPADDING",(0,0),(0,0),0),
                ("TOPPADDING",(0,0),(-1,-1),10*mm),("BOTTOMPADDING",(0,0),(-1,-1),6*mm),
            ]))
        elif header.header_layout=="minimal":
            content_items=[]
            if logo: content_items.append(logo)
            content_items.extend(header_text)
            content_stack=cv_vertical_stack(content_items,page_width-(2*margin),"LEFT")
            modern_header=Table([[content_stack]],colWidths=[page_width])
            modern_header.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,-1),header_bg),("VALIGN",(0,0),(-1,-1),"TOP"),
                ("LEFTPADDING",(0,0),(-1,-1),margin),("RIGHTPADDING",(0,0),(-1,-1),margin),
                ("TOPPADDING",(0,0),(-1,-1),10*mm),("BOTTOMPADDING",(0,0),(-1,-1),6*mm),
            ]))
        else:
            content_items=[]
            if logo: content_items.append(logo)
            content_items.extend(header_text)
            if photo:
                content_stack=cv_vertical_stack(content_items,page_width-photo_col-margin-4*mm,"LEFT")
                modern_header=Table([[content_stack,photo]],colWidths=[page_width-photo_col,photo_col])
                modern_header.setStyle(TableStyle([
                    ("BACKGROUND",(0,0),(-1,-1),header_bg),("VALIGN",(0,0),(-1,-1),"TOP"),
                    ("LEFTPADDING",(0,0),(0,0),margin),("RIGHTPADDING",(1,0),(1,0),margin),
                    ("LEFTPADDING",(1,0),(1,0),0),("RIGHTPADDING",(0,0),(0,0),4*mm),
                    ("ALIGN",(1,0),(1,0),"RIGHT"),
                    ("TOPPADDING",(0,0),(-1,-1),10*mm),("BOTTOMPADDING",(0,0),(-1,-1),6*mm),
                ]))
            else:
                content_stack=cv_vertical_stack(content_items,page_width-(2*margin),"LEFT")
                modern_header=Table([[content_stack]],colWidths=[page_width])
                modern_header.setStyle(TableStyle([
                    ("BACKGROUND",(0,0),(-1,-1),header_bg),("VALIGN",(0,0),(-1,-1),"TOP"),
                    ("LEFTPADDING",(0,0),(-1,-1),margin),("RIGHTPADDING",(0,0),(-1,-1),margin),
                    ("TOPPADDING",(0,0),(-1,-1),10*mm),("BOTTOMPADDING",(0,0),(-1,-1),6*mm),
                ]))
        story.append(modern_header)
        story.append(FrameBreak())
    elif header.header_layout=="centered":
        centered=[]
        if logo: centered.append(logo)
        centered.extend(header_text)
        if photo: centered.append(photo)
        stack=cv_vertical_stack(centered,layout_width,"CENTER")
        wrap=Table([[stack]],colWidths=[layout_width])
        wrap.setStyle(TableStyle([
            ("ALIGN",(0,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0),
            ("TOPPADDING",(0,0),(-1,-1),0),("BOTTOMPADDING",(0,0),(-1,-1),0),
        ]))
        story.append(wrap)
        story.append(Spacer(1,3*mm))
    elif header.header_layout=="minimal":
        minimal=[]
        if logo: minimal.append(logo)
        minimal.extend(header_text)
        story.append(cv_vertical_stack(minimal,layout_width,"LEFT"))
        story.append(Spacer(1,3*mm))
    elif header.header_layout=="photo_left" and photo:
        content_items=[]
        if logo: content_items.append(logo)
        content_items.extend(header_text)
        photo_width=min(35*mm,layout_width*0.22)
        content_stack=cv_vertical_stack(content_items,layout_width-photo_width-5*mm,"LEFT")
        wrap=Table([[photo,content_stack]],colWidths=[photo_width,layout_width-photo_width])
        wrap.setStyle(TableStyle([
            ("VALIGN",(0,0),(-1,-1),"TOP"),
            ("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(0,0),5),
            ("RIGHTPADDING",(1,0),(1,0),0),
            ("TOPPADDING",(0,0),(-1,-1),0),("BOTTOMPADDING",(0,0),(-1,-1),0),
        ]))
        story.append(wrap)
        story.append(Spacer(1,3*mm))
    else:
        left_items=[]
        if logo: left_items.append(logo)
        left_items.extend(header_text)
        if photo:
            photo_width=min(35*mm,layout_width*0.22)
            left_stack=cv_vertical_stack(left_items,layout_width-photo_width-5*mm,"LEFT")
            wrap=Table([[left_stack,photo]],colWidths=[layout_width-photo_width,photo_width])
            wrap.setStyle(TableStyle([
                ("VALIGN",(0,0),(-1,-1),"TOP"),
                ("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(0,0),5),
                ("RIGHTPADDING",(1,0),(1,0),0),("ALIGN",(1,0),(1,0),"RIGHT"),
                ("TOPPADDING",(0,0),(-1,-1),0),("BOTTOMPADDING",(0,0),(-1,-1),0),
            ]))
            story.append(wrap)
        else:
            story.append(cv_vertical_stack(left_items,layout_width,"LEFT"))
        story.append(Spacer(1,3*mm))

    summary=application_cv.profile_summary or profile.get("summary")
    if summary:
        story.extend([Paragraph("Profil",styles["CVSection"]),Paragraph(pdf_text(summary),styles["CVBody"])])

    def row_title(section,d):
        if section=="experience":
            return d.get("position") or "", d.get("employer") or ""
        if section=="education":
            return d.get("degree") or d.get("field_of_study") or d.get("institution") or "", d.get("institution") or ""
        if section=="skills":
            return d.get("name") or "", " · ".join(x for x in [d.get("category"),d.get("level")] if x)
        if section=="languages":
            return d.get("language") or "", d.get("level") or ""
        if section=="certifications":
            return d.get("name") or "", d.get("issuer") or ""
        if section=="projects":
            return d.get("name") or "", " · ".join(x for x in [d.get("role"),d.get("technologies")] if x)
        return d.get("title") or "", " · ".join(x for x in [d.get("category"),d.get("organization")] if x)

    if template_slug=="modern":
        skill_entries=groups.get("skills",[]) if "skills" in section_order else []
        language_entries=groups.get("languages",[]) if "languages" in section_order else []

        def draw_modern_sidebar(canvas,doc_obj):
            # The sidebar is deliberately outside the story flow. This means
            # it occupies the right side of page 1 while the main CV starts
            # at the same vertical position in the left frame.
            if canvas.getPageNumber()!=1:
                return
            canvas.saveState()
            x=margin+main_width+sidebar_gap
            top=page_height-metrics["modern_header_mm"]*mm-5*mm
            bottom=margin
            canvas.setStrokeColor(colors.HexColor("#d0d5dd"))
            canvas.setLineWidth(0.5)
            canvas.line(x-sidebar_gap/2,bottom,x-sidebar_gap/2,top)
            y=top
            available=sidebar_width

            def draw_flow(flow,space_after=1.5*mm):
                nonlocal y
                w,h=flow.wrap(available,max(0,y-bottom))
                if y-h < bottom:
                    return False
                flow.drawOn(canvas,x,y-h)
                y-=h+space_after
                return True

            if skill_entries:
                draw_flow(Paragraph("Kenntnisse & Skills",styles["CVSection"]),2*mm)
                for entry in skill_entries:
                    d=application_cv_entry_data(entry); title,sub=row_title("skills",d)
                    value=f"<b>{escape(title)}</b>"
                    if sub: value += f"<br/><font size='7.5'>{escape(sub)}</font>"
                    if not draw_flow(Paragraph(value,styles["CVBody"]),1.5*mm):
                        draw_flow(Paragraph("Weitere Skills siehe CV-Daten.",styles["CVSmall"]),0)
                        break
            if language_entries and y-bottom > 22*mm:
                y-=2*mm
                draw_flow(Paragraph("Sprachen",styles["CVSection"]),2*mm)
                for entry in language_entries:
                    d=application_cv_entry_data(entry); title,sub=row_title("languages",d)
                    value=f"<b>{escape(title)}</b>"
                    if sub: value += f"<br/><font size='7.5'>{escape(sub)}</font>"
                    if not draw_flow(Paragraph(value,styles["CVBody"]),1.5*mm):
                        break
            canvas.restoreState()

        header_h=metrics["modern_header_mm"]*mm
        content_top=page_height-header_h-5*mm
        header_frame=Frame(
            0,page_height-header_h,page_width,header_h,
            id="modern-header",leftPadding=0,rightPadding=0,topPadding=0,bottomPadding=0,
        )
        first_main_frame=Frame(
            margin,margin,main_width,content_top-margin,
            id="modern-main-first",leftPadding=0,rightPadding=0,topPadding=0,bottomPadding=0,
        )
        later_main_frame=Frame(
            margin,margin,main_width,page_height-(2*margin),
            id="modern-main-later",leftPadding=0,rightPadding=0,topPadding=0,bottomPadding=0,
        )
        modern_first=PageTemplate(id="ModernFirst",frames=[header_frame,first_main_frame],onPage=draw_modern_sidebar,autoNextPageTemplate="ModernLater")
        modern_later=PageTemplate(id="ModernLater",frames=[later_main_frame])
        doc.addPageTemplates([modern_first,modern_later])

        # Main content starts on page 1 beside the sidebar and remains in the
        # left content column on following pages. There is no table cell that
        # can force the entire main column onto page 2.
        main_sections=[s for s in section_order if s not in ("skills","languages")]
        for section in main_sections:
            entries=groups.get(section,[])
            if not entries:
                continue
            story.append(Paragraph(CV_SECTION_LABELS[section],styles["CVSection"]))
            for entry in entries:
                d=application_cv_entry_data(entry); title,sub=row_title(section,d); period=cv_period(d)
                period_width=min(27*mm,main_width*0.24)
                title_text=f"<b>{escape(title)}</b>"
                if sub:
                    title_text += f"<br/><font size='8'>{escape(sub)}</font>"
                item_head=Table(
                    [[Paragraph(title_text,styles["CVBody"]),Paragraph(escape(period),styles["CVSmall"])]],
                    colWidths=[main_width-period_width,period_width],
                    hAlign="LEFT"
                )
                item_head.setStyle(TableStyle([
                    ("VALIGN",(0,0),(-1,-1),"TOP"),
                    ("LEFTPADDING",(0,0),(-1,-1),0),
                    ("RIGHTPADDING",(0,0),(-1,-1),0),
                    ("TOPPADDING",(0,0),(-1,-1),0),
                    ("BOTTOMPADDING",(0,0),(-1,-1),2),
                ]))
                story.append(item_head)
                if d.get("description"):
                    story.append(Paragraph(pdf_text(d.get("description")),styles["CVBody"]))
                story.append(Spacer(1,2*mm))
    else:
        for section in section_order:
            entries=groups.get(section,[])
            if not entries: continue
            story.append(Paragraph(CV_SECTION_LABELS[section],styles["CVSection"]))
            if section in ("skills","languages") and template_slug=="compact":
                items=[]
                for entry in entries:
                    d=application_cv_entry_data(entry); title,sub=row_title(section,d)
                    items.append(f"{title}{' ('+sub+')' if sub else ''}")
                story.append(Paragraph(escape(" · ".join(items)),styles["CVBody"]))
                continue
            for entry in entries:
                d=application_cv_entry_data(entry); title,sub=row_title(section,d); period=cv_period(d)
                left=f"<b>{escape(title)}</b>"
                if sub: left += f"<br/><font size='8'>{escape(sub)}</font>"
                right=escape(period)
                period_width=min(30*mm,usable_width*0.20)
                t=Table([[Paragraph(left,styles["CVBody"]),Paragraph(right,styles["CVSmall"])]],colWidths=[usable_width-period_width,period_width])
                t.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)]))
                story.append(t)
                if d.get("description"): story.append(Paragraph(pdf_text(d.get("description")),styles["CVBody"]))
                story.append(Spacer(1,2*mm))

    footer=f"© {date.today().year} {settings.get('copyright_holder','Lange-IT.com')} · {settings.get('author','Peter Lange')}"
    story.extend([Spacer(1,4*mm),Paragraph(escape(footer),styles["CVFooter"])])
    doc.build(story)
    buffer.seek(0)
    return buffer


CV_SECTIONS = {
    "experience": {"title":"Berufserfahrung","model":CVExperience,"fields":[("employer","Arbeitgeber","text",True),("position","Position","text",True),("location","Ort","text",False),("start_date","Von","date",False),("end_date","Bis","date",False),("is_current","Aktuell","checkbox",False),("description","Aufgaben / Erfolge","textarea",False),("sort_order","Reihenfolge","number",False)]},
    "education": {"title":"Ausbildung & Studium","model":CVEducation,"fields":[("institution","Institution","text",True),("degree","Abschluss","text",False),("field_of_study","Fachrichtung","text",False),("location","Ort","text",False),("start_date","Von","date",False),("end_date","Bis","date",False),("is_current","Aktuell","checkbox",False),("description","Details","textarea",False),("sort_order","Reihenfolge","number",False)]},
    "skills": {"title":"Kenntnisse & Skills","model":CVSkill,"fields":[("name","Skill","text",True),("category","Kategorie","text",False),("level","Niveau","text",False),("sort_order","Reihenfolge","number",False)]},
    "languages": {"title":"Sprachen","model":CVLanguage,"fields":[("language","Sprache","text",True),("level","Niveau","text",False),("sort_order","Reihenfolge","number",False)]},
    "certifications": {"title":"Zertifikate & Weiterbildungen","model":CVCertification,"fields":[("name","Bezeichnung","text",True),("issuer","Aussteller / Anbieter","text",False),("issue_date","Ausgestellt am","date",False),("expiry_date","Gültig bis","date",False),("credential_id","Zertifikats-ID","text",False),("credential_url","Nachweis-URL","url",False),("description","Details","textarea",False),("sort_order","Reihenfolge","number",False)]},
    "projects": {"title":"Projekte","model":CVProject,"fields":[("name","Projekt","text",True),("role","Rolle","text",False),("start_date","Von","date",False),("end_date","Bis","date",False),("url","Projekt-URL","url",False),("technologies","Technologien","text",False),("description","Beschreibung / Ergebnis","textarea",False),("sort_order","Reihenfolge","number",False)]},
    "other": {"title":"Weitere Stationen","model":CVOther,"fields":[("category","Kategorie","text",False),("title","Bezeichnung","text",True),("organization","Organisation","text",False),("location","Ort","text",False),("start_date","Von","date",False),("end_date","Bis","date",False),("description","Details","textarea",False),("sort_order","Reihenfolge","number",False)]}
}

def cv_section_or_404(section):
    from flask import abort
    cfg=CV_SECTIONS.get(section)
    if not cfg: abort(404)
    return cfg

def cv_save_item(item, fields):
    for name, label, field_type, required in fields:
        if field_type == "checkbox":
            value = request.form.get(name) == "on"
        elif field_type == "date":
            value = parse_date(request.form.get(name))
        elif field_type == "number":
            raw=request.form.get(name,"").strip(); value=int(raw) if raw else 0
        else:
            value=request.form.get(name,"").strip() or None
        if required and not value:
            raise ValueError(f"{label} ist erforderlich")
        setattr(item,name,value)


def create_app():
    app=Flask(__name__)
    app.config.update(SECRET_KEY=os.getenv("SECRET_KEY","dev-secret"),SQLALCHEMY_DATABASE_URI=os.getenv("DATABASE_URL","mysql+pymysql://bewerbung:bewerbung@db:3306/bewerbungen?charset=utf8mb4"),SQLALCHEMY_TRACK_MODIFICATIONS=False,MAX_CONTENT_LENGTH=int(os.getenv("MAX_CONTENT_LENGTH",str(20*1024*1024))),UPLOAD_DIR=os.getenv("UPLOAD_DIR","/data/attachments"))
    Path(app.config["UPLOAD_DIR"]).mkdir(parents=True,exist_ok=True); db.init_app(app)
    with app.app_context():
        db.create_all()
        if Status.query.count()==0:
            for i,n in enumerate(DEFAULT_STATUSES): db.session.add(Status(name=n,sort_order=i))
            db.session.commit()
        ensure_default_settings()
        ensure_default_cv_templates()
        ensure_cv_template_styles()
        ensure_cover_letter_defaults()
        ensure_cover_letter_cover_defaults()
        ensure_cover_letter_block_defaults()
        ensure_cover_letter_footer_defaults()
        ensure_status_classifications()

    @app.context_processor
    def globals_(): return {"today":date.today(),"statuses":get_statuses(),"reminder_count":len(reminder_items()),"branding":get_app_settings(),"menu_groups":grouped_menu(),"app_version":APP_VERSION}

    @app.get("/")
    def index():
        query, filters = application_filter_query()
        page=request.args.get("page",1,type=int)
        pg=query.paginate(page=page,per_page=15,error_out=False)
        return render_template(
            "index.html",
            applications=pg.items,
            pagination=pg,
            q=filters["q"],
            selected_status=filters["status"],
            sort=filters["sort"],
            date_from=filters["date_from"],
            date_to=filters["date_to"],
            process_mode=filters["process_mode"],
            total=Application.query.count(),
            active=active_application_count(),
            interviews=Application.query.join(Status).filter(Status.name.in_(["Telefoninterview","Vorstellungsgespräch","Zweitgespräch"])).count(),
            offers=Application.query.join(Status).filter(Status.name=="Angebot").count()
        )

    @app.get("/dashboard")
    def dashboard():
        counts=[{"name":s.name,"count":Application.query.filter_by(status_id=s.id).count()} for s in get_statuses()]
        upcoming=Application.query.filter(Application.follow_up_date.is_not(None),Application.follow_up_date>=date.today()).order_by(Application.follow_up_date).limit(10).all(); overdue=Application.query.filter(Application.follow_up_date.is_not(None),Application.follow_up_date<date.today()).order_by(Application.follow_up_date).limit(10).all()
        return render_template("dashboard.html",total=Application.query.count(),active=active_application_count(),interviews=Application.query.join(Status).filter(Status.name.in_(["Telefoninterview","Vorstellungsgespräch","Zweitgespräch"])).count(),offers=Application.query.join(Status).filter(Status.name=="Angebot").count(),status_counts=counts,upcoming=upcoming,overdue=overdue)

    @app.get("/calendar")
    def calendar_view():
        today=date.today(); year=request.args.get("year",today.year,type=int); month=request.args.get("month",today.month,type=int)
        if month<1: year-=1; month=12
        if month>12: year+=1; month=1
        cal=calendar.Calendar(firstweekday=0); weeks=cal.monthdatescalendar(year,month); events={}
        for a in Application.query.filter(Application.follow_up_date.is_not(None)).all(): events.setdefault(a.follow_up_date,[]).append(("followup",a))
        for a in Application.query.join(ApplicationDetail).filter(ApplicationDetail.interview_date.is_not(None)).all(): events.setdefault(a.details.interview_date,[]).append(("interview",a))
        prev_m=month-1 or 12; prev_y=year-1 if month==1 else year; next_m=month+1 if month<12 else 1; next_y=year+1 if month==12 else year
        return render_template("calendar.html",weeks=weeks,events=events,year=year,month=month,month_name=["","Januar","Februar","März","April","Mai","Juni","Juli","August","September","Oktober","November","Dezember"][month],prev_y=prev_y,prev_m=prev_m,next_y=next_y,next_m=next_m)

    @app.get("/reminders")
    def reminders(): return render_template("reminders.html",items=reminder_items())

    @app.get("/kanban")
    def kanban(): return render_template("kanban.html",columns=[{"status":s,"applications":Application.query.filter_by(status_id=s.id).order_by(Application.updated_at.desc()).all()} for s in get_statuses()])

    @app.post("/api/applications/<int:application_id>/status")
    def kanban_status(application_id):
        a=Application.query.get_or_404(application_id); data=request.get_json(silent=True) or {}; s=Status.query.get(data.get("status_id"))
        if not s:return jsonify(ok=False,error="Status nicht gefunden"),404
        old=a.status.name
        if old!=s.name: a.status=s; log_event(a,"status",f"Status geändert: „{old}“ → „{s.name}“"); db.session.commit()
        return jsonify(ok=True,status=s.name)

    @app.get("/export.xlsx")
    def export_excel():
        query,filters=application_filter_query()
        wb=Workbook(); ws=wb.active; ws.title="Bewerbungen"; headers=["Referenz","Unternehmen","Ansprechpartner","Position","Standort","Bewerbungsdatum","Status","Nächste Aktion","Wiedervorlage","Quelle","Beschäftigungsart","Arbeitsmodell","Priorität","Gesprächsdatum","Gesprächsort","Gehaltsvorstellung","Angebotenes Gehalt","Absagegrund","Stellenausschreibung","E-Mail","Telefon","Notizen","Erstellt","Geändert"]; ws.append(headers)
        for a in query.all():
            d=a.details; ws.append([a.reference,a.company,a.contact_person,a.position,a.location,a.application_date,a.status.name,a.next_action,a.follow_up_date,d.source if d else None,d.employment_type if d else None,d.work_model if d else None,d.priority if d else None,d.interview_date if d else None,d.interview_location if d else None,a.salary_expectation,d.offered_salary if d else None,d.rejection_reason if d else None,a.job_url,a.email,a.phone,a.notes,a.created_at,a.updated_at])
        ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
        for i in range(1,len(headers)+1): ws.column_dimensions[get_column_letter(i)].width=22
        mode_suffix={"active":"aktiv","completed":"erledigt","all":"alle"}.get(filters["process_mode"],"gefiltert")
        path="/tmp/bewerbungen.xlsx"; wb.save(path); return send_file(path,as_attachment=True,download_name=f"bewerbungen_{mode_suffix}_{date.today().isoformat()}.xlsx",mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    def save_form(a):
        old=a.status.name if a.id else None
        a.company=request.form["company"].strip(); a.position=request.form["position"].strip(); a.contact_person=request.form.get("contact_person","").strip() or None; a.location=request.form.get("location","").strip() or None; a.application_date=parse_date(request.form.get("application_date")); a.status=Status.query.get_or_404(request.form.get("status_id",type=int)); a.next_action=request.form.get("next_action","").strip() or None; a.follow_up_date=parse_date(request.form.get("follow_up_date")); a.job_url=request.form.get("job_url","").strip() or None; a.email=request.form.get("email","").strip() or None; a.phone=request.form.get("phone","").strip() or None; a.salary_expectation=request.form.get("salary_expectation","").strip() or None; a.notes=request.form.get("notes","").strip() or None
        d=detail_for(a); d.source=request.form.get("source","").strip() or None; d.employment_type=request.form.get("employment_type","").strip() or None; d.work_model=request.form.get("work_model","").strip() or None; d.priority=request.form.get("priority","").strip() or "Normal"; d.interview_date=parse_date(request.form.get("interview_date")); d.interview_location=request.form.get("interview_location","").strip() or None; d.offered_salary=request.form.get("offered_salary","").strip() or None; d.rejection_reason=request.form.get("rejection_reason","").strip() or None; d.reminder_enabled=request.form.get("reminder_enabled")=="on"; d.reminder_days_before=max(0,min(30,request.form.get("reminder_days_before",2,type=int)))
        cd=company_detail_for(a); cd.department=request.form.get("company_department","").strip() or None; cd.street=request.form.get("company_street","").strip() or None; cd.postal_code=request.form.get("company_postal_code","").strip() or None; cd.city=request.form.get("company_city","").strip() or None; cd.country=request.form.get("company_country","").strip() or None
        return old

    @app.route("/applications/new",methods=["GET","POST"])
    def new_application():
        if request.method=="POST":
            try:
                a=Application(reference=make_reference(),company="",position="",status=get_statuses()[0]); db.session.add(a); old=save_form(a); db.session.flush(); log_event(a,"created","Bewerbungsvorgang angelegt."); db.session.commit(); flash("Bewerbung wurde angelegt.","success"); return redirect(url_for("edit_application",application_id=a.id))
            except (ValueError,KeyError): db.session.rollback(); flash("Bitte prüfe die eingegebenen Daten.","danger")
        return render_template("form.html",application=None,page_title="Neue Bewerbung")

    @app.route("/applications/<int:application_id>/edit",methods=["GET","POST"])
    def edit_application(application_id):
        a=Application.query.get_or_404(application_id)
        if request.method=="POST":
            old=save_form(a)
            if old!=a.status.name: log_event(a,"status",f"Status geändert: „{old}“ → „{a.status.name}“")
            else: log_event(a,"updated","Bewerbungsvorgang bearbeitet.")
            db.session.commit(); flash("Bewerbung wurde gespeichert.","success"); return redirect(url_for("edit_application",application_id=a.id))
        return render_template("form.html",application=a,page_title=f"Bewerbung {a.reference}")




    @app.get("/cv")
    def cv_index():
        profile=CVProfile.query.first()
        data={}
        for key,cfg in CV_SECTIONS.items():
            model=cfg["model"]
            # explicit stable ordering, then newest dates where available
            if hasattr(model,"start_date"):
                items=model.query.order_by(model.sort_order.asc(), model.start_date.desc(), model.id.desc()).all()
            else:
                items=model.query.order_by(model.sort_order.asc(), model.id.asc()).all()
            data[key]=items
        return render_template("cv.html", profile=profile, sections=CV_SECTIONS, data=data)

    @app.route("/cv/profile", methods=["GET","POST"])
    def cv_profile():
        profile=CVProfile.query.first()
        if profile is None:
            profile=CVProfile(); db.session.add(profile)
        if request.method=="POST":
            action=request.form.get("action","save")
            if action=="remove_photo":
                old_photo=master_cv_photo_filename()
                if old_photo: safe_remove_cv_media(old_photo)
                set_app_setting("cv_profile_photo_filename","")
                db.session.commit()
                flash("Master-Profilfoto entfernt.","success")
                return redirect(url_for("cv_profile"))

            profile.first_name=request.form.get("first_name","").strip() or None; profile.last_name=request.form.get("last_name","").strip() or None; profile.professional_title=request.form.get("professional_title","").strip() or None
            profile.birth_date=parse_date(request.form.get("birth_date")); profile.birth_place=request.form.get("birth_place","").strip() or None; profile.address=request.form.get("address","").strip() or None; profile.postal_code=request.form.get("postal_code","").strip() or None; profile.city=request.form.get("city","").strip() or None; profile.country=request.form.get("country","").strip() or None
            profile.email=request.form.get("email","").strip() or None; profile.phone=request.form.get("phone","").strip() or None; profile.website=request.form.get("website","").strip() or None; profile.linkedin=request.form.get("linkedin","").strip() or None; profile.xing=request.form.get("xing","").strip() or None; profile.github=request.form.get("github","").strip() or None; profile.summary=request.form.get("summary","").strip() or None

            photo=request.files.get("profile_photo")
            if photo and photo.filename:
                try:
                    old_photo=master_cv_photo_filename()
                    new_photo=save_cv_media_upload(photo,"master","profile")
                    if old_photo: safe_remove_cv_media(old_photo)
                    set_app_setting("cv_profile_photo_filename",new_photo)
                except ValueError as exc:
                    flash(str(exc),"danger")
                    return render_template("cv_profile.html",profile=profile,master_photo=master_cv_photo_filename())

            db.session.commit(); flash("Lebenslauf-Stammdaten gespeichert.","success"); return redirect(url_for("cv_index"))
        return render_template("cv_profile.html", profile=profile, master_photo=master_cv_photo_filename())

    @app.route("/cv/<section>/new", methods=["GET","POST"])
    def cv_new(section):
        cfg=cv_section_or_404(section); item=cfg["model"]()
        if request.method=="POST":
            try:
                cv_save_item(item,cfg["fields"]); db.session.add(item); db.session.commit(); flash(f"{cfg['title']}: Eintrag angelegt.","success"); return redirect(url_for("cv_index"))
            except (ValueError, TypeError) as exc:
                db.session.rollback(); flash(str(exc) or "Bitte prüfe die Eingaben.","danger")
        return render_template("cv_form.html", section=section, cfg=cfg, item=None)

    @app.route("/cv/<section>/<int:item_id>/edit", methods=["GET","POST"])
    def cv_edit(section,item_id):
        cfg=cv_section_or_404(section); item=cfg["model"].query.get_or_404(item_id)
        if request.method=="POST":
            try:
                cv_save_item(item,cfg["fields"]); db.session.commit(); flash(f"{cfg['title']}: Eintrag gespeichert.","success"); return redirect(url_for("cv_index"))
            except (ValueError, TypeError) as exc:
                db.session.rollback(); flash(str(exc) or "Bitte prüfe die Eingaben.","danger")
        return render_template("cv_form.html", section=section, cfg=cfg, item=item)

    @app.post("/cv/<section>/<int:item_id>/delete")
    def cv_delete(section,item_id):
        cfg=cv_section_or_404(section); item=cfg["model"].query.get_or_404(item_id); db.session.delete(item); db.session.commit(); flash(f"{cfg['title']}: Eintrag gelöscht.","success"); return redirect(url_for("cv_index"))

    @app.get("/cv/templates")
    def cv_templates():
        templates=CVTemplate.query.order_by(CVTemplate.sort_order.asc(),CVTemplate.name.asc()).all()
        return render_template("cv_templates.html",templates=templates)

    @app.route("/cv/templates/new",methods=["GET","POST"])
    def cv_template_new():
        if request.method=="POST":
            name=request.form.get("name","").strip()
            if not name:
                flash("Name ist erforderlich.","danger")
                return render_template("cv_template_edit.html",template=None,style=None,new_template=True)
            slug_base=re.sub(r"[^a-z0-9]+","-",name.lower()).strip("-") or "template"
            slug=slug_base
            i=2
            while CVTemplate.query.filter_by(slug=slug).first():
                slug=f"{slug_base}-{i}"; i+=1
            template=CVTemplate(name=name,slug=slug,description=request.form.get("description","").strip() or None,sort_order=request.form.get("sort_order",100,type=int),is_active=request.form.get("is_active")=="on")
            db.session.add(template); db.session.flush()
            style=CVTemplateStyle(template=template)
            db.session.add(style)
            save_cv_template_style(style)
            db.session.commit()
            flash("Eigenes CV-Template erstellt.","success")
            return redirect(url_for("cv_template_edit",template_id=template.id))
        return render_template("cv_template_edit.html",template=None,style=None,new_template=True)

    def save_cv_template_style(style):
        base=request.form.get("base_layout","classic")
        style.base_layout=base if base in ("classic","modern","compact") else "classic"
        accent=request.form.get("accent_color","#172033").strip()
        style.accent_color=accent if re.match(r"^#[0-9A-Fa-f]{6}$",accent) else "#172033"
        font=request.form.get("font_family","Helvetica")
        style.font_family=font if font in ("Helvetica","Times-Roman","Courier") else "Helvetica"
        try: scale=float(request.form.get("font_scale","1.0"))
        except (TypeError,ValueError): scale=1.0
        style.font_scale=max(0.8,min(1.3,scale))
        style.page_margin_mm=max(8,min(25,request.form.get("page_margin_mm",15,type=int)))
        style.show_logo=request.form.get("show_logo")=="on"
        style.show_company=request.form.get("show_company")=="on"
        style.show_target_position=request.form.get("show_target_position")=="on"
        order=request.form.get("section_order","experience,education,projects,skills,languages,certifications,other")
        valid=[x.strip() for x in order.split(",") if x.strip() in CV_SECTION_LABELS]
        for key in CV_SECTION_LABELS:
            if key not in valid: valid.append(key)
        style.section_order=",".join(valid)

    @app.route("/cv/templates/<int:template_id>/edit",methods=["GET","POST"])
    def cv_template_edit(template_id):
        template=CVTemplate.query.get_or_404(template_id)
        style=template.style
        if not style:
            style=CVTemplateStyle(template=template); db.session.add(style); db.session.flush()
        if request.method=="POST":
            template.name=request.form.get("name","").strip() or template.name
            template.description=request.form.get("description","").strip() or None
            template.is_active=request.form.get("is_active")=="on"
            template.sort_order=request.form.get("sort_order",0,type=int)
            save_cv_template_style(style)
            db.session.commit()
            flash("CV-Template und Design gespeichert.","success")
            return redirect(url_for("cv_template_edit",template_id=template.id))
        return render_template("cv_template_edit.html",template=template,style=style,new_template=False)

    @app.post("/cv/templates/<int:template_id>/duplicate")
    def cv_template_duplicate(template_id):
        source=CVTemplate.query.get_or_404(template_id)
        name=f"{source.name} Kopie"
        slug_base=f"{source.slug}-copy"; slug=slug_base; i=2
        while CVTemplate.query.filter_by(slug=slug).first():
            slug=f"{slug_base}-{i}"; i+=1
        new_template=CVTemplate(name=name,slug=slug,description=source.description,sort_order=source.sort_order+1,is_active=True)
        db.session.add(new_template); db.session.flush()
        s=template_style(source)
        db.session.add(CVTemplateStyle(
            template=new_template,base_layout=s.base_layout,accent_color=s.accent_color,font_family=s.font_family,
            font_scale=s.font_scale,page_margin_mm=s.page_margin_mm,show_logo=s.show_logo,
            show_company=s.show_company,show_target_position=s.show_target_position,section_order=s.section_order
        ))
        db.session.commit()
        flash("Template wurde dupliziert und kann jetzt angepasst werden.","success")
        return redirect(url_for("cv_template_edit",template_id=new_template.id))

    @app.route("/applications/<int:application_id>/cvs/new",methods=["GET","POST"])
    def application_cv_new(application_id):
        application=Application.query.get_or_404(application_id)
        profile=CVProfile.query.first()
        templates=CVTemplate.query.filter_by(is_active=True).order_by(CVTemplate.sort_order.asc()).all()
        universal_options=UniversalCV.query.order_by(UniversalCV.updated_at.desc()).all()
        if not templates:
            ensure_default_cv_templates()
            templates=CVTemplate.query.filter_by(is_active=True).order_by(CVTemplate.sort_order.asc()).all()
        if request.method=="POST":
            source_mode=request.form.get("source_mode","master")
            try:
                if source_mode=="universal":
                    universal=UniversalCV.query.get_or_404(request.form.get("universal_cv_id",type=int))
                    cv=copy_universal_cv_to_application(universal,application)
                    cv.title=request.form.get("title",cv.title).strip() or cv.title
                    cv.target_company=request.form.get("target_company",application.company).strip() or application.company
                    cv.target_position=request.form.get("target_position",application.position).strip() or application.position
                    if request.form.get("profile_summary","").strip():
                        cv.profile_summary=request.form.get("profile_summary","").strip()
                else:
                    template=CVTemplate.query.get_or_404(request.form.get("template_id",type=int))
                    cv=create_application_cv_snapshot(
                        application=application,template=template,
                        title=request.form.get("title","Lebenslauf"),
                        target_company=request.form.get("target_company",""),
                        target_position=request.form.get("target_position",""),
                        profile_summary=request.form.get("profile_summary",""),
                    )
                    configure_new_cv_header(cv)
                log_event(application,"cv",f"Lebenslauf-Snapshot erstellt: {cv.title}")
                db.session.commit()
            except ValueError as exc:
                db.session.rollback(); flash(str(exc),"danger")
                return render_template("application_cv_new.html",application=application,templates=templates,profile=profile,master_photo=master_cv_photo_filename(),has_system_logo=bool(get_app_settings().get("logo_filename")),universal_options=universal_options)
            flash("Bewerbungsspezifischer Lebenslauf wurde erstellt.","success")
            return redirect(url_for("application_cv_edit",cv_id=cv.id))
        return render_template("application_cv_new.html",application=application,templates=templates,profile=profile,master_photo=master_cv_photo_filename(),has_system_logo=bool(get_app_settings().get("logo_filename")),universal_options=universal_options)

    @app.route("/application-cvs/<int:cv_id>/edit",methods=["GET","POST"])
    def application_cv_edit(cv_id):
        cv=ApplicationCV.query.get_or_404(cv_id)
        header=ensure_application_cv_header(cv)
        templates=CVTemplate.query.order_by(CVTemplate.sort_order.asc()).all()
        if request.method=="POST":
            action=request.form.get("action","save")
            if action=="remove_header_logo":
                safe_remove_cv_media(header.logo_filename); header.logo_filename=None
                db.session.commit(); flash("CV-Logo aus diesem Snapshot entfernt.","success")
                return redirect(url_for("application_cv_edit",cv_id=cv.id))
            if action=="remove_header_photo":
                safe_remove_cv_media(header.photo_filename); header.photo_filename=None
                db.session.commit(); flash("CV-Foto aus diesem Snapshot entfernt.","success")
                return redirect(url_for("application_cv_edit",cv_id=cv.id))
            if action=="copy_system_logo":
                safe_remove_cv_media(header.logo_filename)
                header.logo_filename=system_logo_relative()
                db.session.commit(); flash("Aktuelles B-V-S-Logo wurde als Snapshot in diesen CV kopiert.","success")
                return redirect(url_for("application_cv_edit",cv_id=cv.id))
            if action=="copy_master_photo":
                safe_remove_cv_media(header.photo_filename)
                header.photo_filename=copy_cv_media(master_cv_photo_filename(),f"snapshots/{cv.id}","photo")
                db.session.commit(); flash("Master-Profilfoto wurde als Snapshot in diesen CV kopiert.","success")
                return redirect(url_for("application_cv_edit",cv_id=cv.id))

            cv.title=request.form.get("title","").strip() or "Lebenslauf"
            cv.target_company=request.form.get("target_company","").strip() or None
            cv.target_position=request.form.get("target_position","").strip() or None
            cv.profile_summary=request.form.get("profile_summary","").strip() or None
            template_id=request.form.get("template_id",type=int)
            if template_id:
                cv.template=CVTemplate.query.get_or_404(template_id)

            header.document_title=request.form.get("document_title","Lebenslauf").strip() or "Lebenslauf"
            header.header_subtitle=request.form.get("header_subtitle","").strip() or None
            layout=request.form.get("header_layout","photo_right")
            header.header_layout=layout if layout in ("photo_right","photo_left","centered","minimal") else "photo_right"
            header.show_document_title=request.form.get("show_document_title")=="on"
            header.show_professional_title=request.form.get("show_professional_title")=="on"
            header.show_target_company=request.form.get("show_target_company")=="on"
            header.show_target_position=request.form.get("show_target_position")=="on"
            header.show_contact=request.form.get("show_contact")=="on"

            try:
                logo=request.files.get("header_logo")
                if logo and logo.filename:
                    new_logo=save_cv_media_upload(logo,f"snapshots/{cv.id}","logo")
                    safe_remove_cv_media(header.logo_filename); header.logo_filename=new_logo
                photo=request.files.get("header_photo")
                if photo and photo.filename:
                    new_photo=save_cv_media_upload(photo,f"snapshots/{cv.id}","photo")
                    safe_remove_cv_media(header.photo_filename); header.photo_filename=new_photo
            except ValueError as exc:
                db.session.rollback(); flash(str(exc),"danger")
                return redirect(url_for("application_cv_edit",cv_id=cv.id))

            for entry in cv.entries:
                entry.visible=request.form.get(f"visible_{entry.id}")=="on"
                entry.sort_order=request.form.get(f"sort_{entry.id}",entry.sort_order,type=int)
            db.session.commit()
            flash("CV-Version und Kopfzeilen-Snapshot gespeichert.","success")
            return redirect(url_for("application_cv_edit",cv_id=cv.id))
        return render_template(
            "application_cv_edit.html",
            cv=cv,header=header,templates=templates,
            groups=cv_entry_groups(cv),section_labels=CV_SECTION_LABELS,
            entry_data=application_cv_entry_data,
            master_photo=master_cv_photo_filename(),
            has_system_logo=bool(get_app_settings().get("logo_filename"))
        )

    @app.route("/application-cv-entries/<int:entry_id>/edit",methods=["GET","POST"])
    def application_cv_entry_edit(entry_id):
        entry=ApplicationCVEntry.query.get_or_404(entry_id)
        cfg=cv_section_or_404(entry.section)
        data=application_cv_entry_data(entry)
        if request.method=="POST":
            updated={}
            for name,label,field_type,required in cfg["fields"]:
                if field_type=="checkbox":
                    value=request.form.get(name)=="on"
                elif field_type=="date":
                    raw=request.form.get(name,"").strip()
                    value=raw or None
                elif field_type=="number":
                    raw=request.form.get(name,"").strip()
                    value=int(raw) if raw else 0
                else:
                    value=request.form.get(name,"").strip() or None
                if required and not value:
                    flash(f"{label} ist erforderlich.","danger")
                    return render_template("application_cv_entry_edit.html",entry=entry,cfg=cfg,data=request.form)
                updated[name]=value
            entry.data_json=json.dumps(updated,ensure_ascii=False)
            db.session.commit()
            flash("Snapshot-Eintrag gespeichert. Der Master-Lebenslauf wurde nicht verändert.","success")
            return redirect(url_for("application_cv_edit",cv_id=entry.application_cv_id))
        return render_template("application_cv_entry_edit.html",entry=entry,cfg=cfg,data=data)

    @app.get("/application-cvs/<int:cv_id>/preview")
    def application_cv_preview(cv_id):
        cv=ApplicationCV.query.get_or_404(cv_id)
        header=ensure_application_cv_header(cv)
        db.session.commit()
        return render_template(
            "application_cv_preview.html",
            cv=cv,header=header,profile=application_cv_profile(cv),
            groups=cv_entry_groups(cv,visible_only=True),
            section_labels=CV_SECTION_LABELS,
            entry_data=application_cv_entry_data,
            period=cv_period,
            style=template_style(cv.template),
            metrics=cv_render_metrics(template_style(cv.template)),
            section_order=normalized_section_order(template_style(cv.template)),
        )

    @app.get("/application-cvs/<int:cv_id>/pdf")
    def application_cv_pdf(cv_id):
        cv=ApplicationCV.query.get_or_404(cv_id)
        try:
            ensure_application_cv_header(cv)
            db.session.commit()
            pdf=build_application_cv_pdf(cv)
            filename=f"{secure_filename(cv.application.reference)}_{secure_filename(cv.target_company or cv.application.company)}_CV.pdf"
            return send_file(pdf,as_attachment=True,download_name=filename,mimetype="application/pdf")
        except Exception as exc:
            db.session.rollback()
            current_app.logger.exception("CV-PDF-Export fehlgeschlagen für CV %s",cv_id)
            flash(f"CV-PDF konnte nicht erstellt werden: {type(exc).__name__}: {exc}","danger")
            return redirect(url_for("application_cv_edit",cv_id=cv_id))

    @app.post("/application-cvs/<int:cv_id>/delete")
    def application_cv_delete(cv_id):
        cv=ApplicationCV.query.get_or_404(cv_id)
        application_id=cv.application_id
        title=cv.title
        media_dir=cv_media_root() / "snapshots" / str(cv.id)
        db.session.delete(cv)
        db.session.commit()
        if media_dir.exists():
            shutil.rmtree(media_dir,ignore_errors=True)
        flash(f"CV-Version „{title}“ gelöscht.","success")
        return redirect(url_for("edit_application",application_id=application_id))

    @app.get("/cv/universal")
    def universal_cvs():
        return render_template("universal_cvs.html",cvs=UniversalCV.query.order_by(UniversalCV.updated_at.desc()).all())

    @app.route("/cv/universal/new",methods=["GET","POST"])
    def universal_cv_new():
        templates=CVTemplate.query.filter_by(is_active=True).order_by(CVTemplate.sort_order.asc()).all()
        profile=CVProfile.query.first()
        if request.method=="POST":
            template=CVTemplate.query.get_or_404(request.form.get("template_id",type=int))
            try:
                cv=create_universal_cv_snapshot(
                    template=template,title=request.form.get("title","Allgemeiner Lebenslauf"),
                    target_position=request.form.get("target_position",""),
                    profile_summary=request.form.get("profile_summary",""),
                )
                configure_new_cv_header(cv)
                cv.target_company=None
                cv.header.show_target_company=False
                db.session.commit()
            except ValueError as exc:
                db.session.rollback(); flash(str(exc),"danger")
                return render_template("universal_cv_new.html",templates=templates,profile=profile,master_photo=master_cv_photo_filename(),has_system_logo=bool(get_app_settings().get("logo_filename")))
            flash("Universeller CV wurde aus dem Master-Lebenslauf erstellt.","success")
            return redirect(url_for("universal_cv_edit",cv_id=cv.id))
        return render_template("universal_cv_new.html",templates=templates,profile=profile,master_photo=master_cv_photo_filename(),has_system_logo=bool(get_app_settings().get("logo_filename")))

    @app.route("/cv/universal/<int:cv_id>/edit",methods=["GET","POST"])
    def universal_cv_edit(cv_id):
        cv=UniversalCV.query.get_or_404(cv_id)
        header=ensure_application_cv_header(cv)
        templates=CVTemplate.query.order_by(CVTemplate.sort_order.asc()).all()
        if request.method=="POST":
            action=request.form.get("action","save")
            if action=="remove_header_logo":
                safe_remove_cv_media(header.logo_filename); header.logo_filename=None
                db.session.commit(); return redirect(url_for("universal_cv_edit",cv_id=cv.id))
            if action=="remove_header_photo":
                safe_remove_cv_media(header.photo_filename); header.photo_filename=None
                db.session.commit(); return redirect(url_for("universal_cv_edit",cv_id=cv.id))
            if action=="copy_master_photo":
                safe_remove_cv_media(header.photo_filename)
                header.photo_filename=copy_cv_media(master_cv_photo_filename(),f"universal/{cv.id}","photo")
                db.session.commit(); return redirect(url_for("universal_cv_edit",cv_id=cv.id))
            if action=="copy_system_logo":
                safe_remove_cv_media(header.logo_filename); header.logo_filename=system_logo_relative()
                db.session.commit(); return redirect(url_for("universal_cv_edit",cv_id=cv.id))
            cv.title=request.form.get("title","").strip() or "Allgemeiner Lebenslauf"
            cv.target_position=request.form.get("target_position","").strip() or None
            cv.target_company=None
            cv.profile_summary=request.form.get("profile_summary","").strip() or None
            tid=request.form.get("template_id",type=int)
            if tid: cv.template=CVTemplate.query.get_or_404(tid)
            header.document_title=request.form.get("document_title","Lebenslauf").strip() or "Lebenslauf"
            header.header_subtitle=request.form.get("header_subtitle","").strip() or None
            layout=request.form.get("header_layout","photo_right")
            header.header_layout=layout if layout in ("photo_right","photo_left","centered","minimal") else "photo_right"
            header.show_document_title=request.form.get("show_document_title")=="on"
            header.show_professional_title=request.form.get("show_professional_title")=="on"
            header.show_target_company=False
            header.show_target_position=request.form.get("show_target_position")=="on"
            header.show_contact=request.form.get("show_contact")=="on"
            try:
                logo=request.files.get("header_logo")
                if logo and logo.filename:
                    new_logo=save_cv_media_upload(logo,f"universal/{cv.id}","logo"); safe_remove_cv_media(header.logo_filename); header.logo_filename=new_logo
                photo=request.files.get("header_photo")
                if photo and photo.filename:
                    new_photo=save_cv_media_upload(photo,f"universal/{cv.id}","photo"); safe_remove_cv_media(header.photo_filename); header.photo_filename=new_photo
            except ValueError as exc:
                db.session.rollback(); flash(str(exc),"danger"); return redirect(url_for("universal_cv_edit",cv_id=cv.id))
            for entry in cv.entries:
                entry.visible=request.form.get(f"visible_{entry.id}")=="on"
                entry.sort_order=request.form.get(f"sort_{entry.id}",entry.sort_order,type=int)
            db.session.commit(); flash("Universeller CV gespeichert.","success")
            return redirect(url_for("universal_cv_edit",cv_id=cv.id))
        return render_template("universal_cv_edit.html",cv=cv,header=header,templates=templates,groups=cv_entry_groups(cv),section_labels=CV_SECTION_LABELS,entry_data=application_cv_entry_data,master_photo=master_cv_photo_filename(),has_system_logo=bool(get_app_settings().get("logo_filename")))

    @app.route("/cv/universal/entries/<int:entry_id>/edit",methods=["GET","POST"])
    def universal_cv_entry_edit(entry_id):
        entry=UniversalCVEntry.query.get_or_404(entry_id)
        cfg=cv_section_or_404(entry.section); data=application_cv_entry_data(entry)
        if request.method=="POST":
            updated={}
            for name,label,field_type,required in cfg["fields"]:
                if field_type=="checkbox": value=request.form.get(name)=="on"
                elif field_type=="date": value=request.form.get(name,"").strip() or None
                elif field_type=="number":
                    raw=request.form.get(name,"").strip(); value=int(raw) if raw else 0
                else: value=request.form.get(name,"").strip() or None
                if required and not value:
                    flash(f"{label} ist erforderlich.","danger")
                    return render_template("universal_cv_entry_edit.html",entry=entry,cfg=cfg,data=request.form)
                updated[name]=value
            entry.data_json=json.dumps(updated,ensure_ascii=False); db.session.commit()
            flash("Universeller CV-Eintrag gespeichert.","success")
            return redirect(url_for("universal_cv_edit",cv_id=entry.universal_cv_id))
        return render_template("universal_cv_entry_edit.html",entry=entry,cfg=cfg,data=data)

    @app.get("/cv/universal/<int:cv_id>/preview")
    def universal_cv_preview(cv_id):
        cv=UniversalCV.query.get_or_404(cv_id); header=ensure_application_cv_header(cv); db.session.commit()
        return render_template("universal_cv_preview.html",cv=cv,header=header,profile=application_cv_profile(cv),groups=cv_entry_groups(cv,visible_only=True),section_labels=CV_SECTION_LABELS,entry_data=application_cv_entry_data,period=cv_period,style=template_style(cv.template),metrics=cv_render_metrics(template_style(cv.template)),section_order=normalized_section_order(template_style(cv.template)))

    @app.get("/cv/universal/<int:cv_id>/pdf")
    def universal_cv_pdf(cv_id):
        cv=UniversalCV.query.get_or_404(cv_id)
        try:
            ensure_application_cv_header(cv); db.session.commit(); pdf=build_application_cv_pdf(cv)
            filename=f"{secure_filename(cv.title)}_CV.pdf"
            return send_file(pdf,as_attachment=True,download_name=filename,mimetype="application/pdf")
        except Exception as exc:
            db.session.rollback(); current_app.logger.exception("Universal-CV-PDF fehlgeschlagen")
            flash(f"CV-PDF konnte nicht erstellt werden: {type(exc).__name__}: {exc}","danger")
            return redirect(url_for("universal_cv_edit",cv_id=cv.id))

    @app.post("/cv/universal/<int:cv_id>/delete")
    def universal_cv_delete(cv_id):
        cv=UniversalCV.query.get_or_404(cv_id); media=cv_media_root()/ "universal" / str(cv.id)
        db.session.delete(cv); db.session.commit(); shutil.rmtree(media,ignore_errors=True)
        flash("Universeller CV gelöscht.","success"); return redirect(url_for("universal_cvs"))

    @app.route("/cover-letter",methods=["GET","POST"])
    def cover_letter_master():
        master=CoverLetterMaster.query.first()
        if not master:
            ensure_cover_letter_defaults(); master=CoverLetterMaster.query.first()
        if request.method=="POST":
            action=request.form.get("action","save")
            if action=="remove_signature":
                old=master_signature_filename()
                if old: safe_remove_cv_media(old)
                set_app_setting("cover_letter_signature_filename",""); db.session.commit(); flash("Master-Unterschrift entfernt.","success"); return redirect(url_for("cover_letter_master"))
            for field in ["subject","salutation","intro_text","body_text","motivation_text","salary_text","closing_text","signoff"]:
                setattr(master,field,request.form.get(field,"").strip() or None)
            signature=request.files.get("signature")
            if signature and signature.filename:
                try:
                    old=master_signature_filename(); new=save_cv_media_upload(signature,"letters/master","signature")
                    if old: safe_remove_cv_media(old)
                    set_app_setting("cover_letter_signature_filename",new)
                except ValueError as exc: flash(str(exc),"danger"); return redirect(url_for("cover_letter_master"))
            db.session.commit(); flash("Master-Anschreiben gespeichert.","success"); return redirect(url_for("cover_letter_master"))
        return render_template("cover_letter_master.html",master=master,signature=master_signature_filename())

    @app.route("/cover-letter/certifications",methods=["GET","POST"])
    def certification_logos():
        if request.method=="POST":
            name=request.form.get("name","").strip(); issuer=request.form.get("issuer","").strip() or None; upload=request.files.get("logo")
            if not name or not upload or not upload.filename:
                flash("Name und Logo-Datei sind erforderlich.","danger"); return redirect(url_for("certification_logos"))
            try: filename=save_cv_media_upload(upload,"certifications","cert")
            except ValueError as exc: flash(str(exc),"danger"); return redirect(url_for("certification_logos"))
            cert=CertificationLogo(name=name,issuer=issuer,filename=filename,sort_order=request.form.get("sort_order",0,type=int),is_active=True)
            db.session.add(cert); db.session.commit(); flash("Zertifizierungslogo gespeichert.","success"); return redirect(url_for("certification_logos"))
        return render_template("certification_logos.html",certifications=CertificationLogo.query.order_by(CertificationLogo.sort_order,CertificationLogo.id).all())

    @app.post("/cover-letter/certifications/<int:cert_id>/update")
    def certification_logo_update(cert_id):
        cert=CertificationLogo.query.get_or_404(cert_id); cert.name=request.form.get("name","").strip() or cert.name; cert.issuer=request.form.get("issuer","").strip() or None; cert.sort_order=request.form.get("sort_order",cert.sort_order,type=int); cert.is_active=request.form.get("is_active")=="on"
        upload=request.files.get("logo")
        if upload and upload.filename:
            try:
                new=save_cv_media_upload(upload,"certifications","cert"); safe_remove_cv_media(cert.filename); cert.filename=new
            except ValueError as exc: flash(str(exc),"danger"); return redirect(url_for("certification_logos"))
        db.session.commit(); flash("Zertifizierungslogo aktualisiert.","success"); return redirect(url_for("certification_logos"))

    @app.post("/cover-letter/certifications/<int:cert_id>/delete")
    def certification_logo_delete(cert_id):
        cert=CertificationLogo.query.get_or_404(cert_id); safe_remove_cv_media(cert.filename); CoverLetterTemplateCertification.query.filter_by(certification_id=cert.id).delete(); db.session.delete(cert); db.session.commit(); flash("Zertifizierungslogo gelöscht.","success"); return redirect(url_for("certification_logos"))

    @app.get("/cover-letter/templates")
    def cover_letter_templates():
        return render_template("cover_letter_templates.html",templates=CoverLetterTemplate.query.order_by(CoverLetterTemplate.sort_order).all())

    @app.route("/cover-letter/templates/<int:template_id>/edit",methods=["GET","POST"])
    def cover_letter_template_edit(template_id):
        t=CoverLetterTemplate.query.get_or_404(template_id)
        footer=footer_style_for(t); cover=cover_style_for(t)
        ensure_cover_letter_block_defaults()
        certifications=CertificationLogo.query.order_by(CertificationLogo.sort_order.asc(),CertificationLogo.id.asc()).all()
        if request.method=="POST":
            action=request.form.get("action","save")
            if action=="add_cover_block":
                max_order=max([b.sort_order for b in t.cover_blocks] or [0])
                block=CoverLetterCoverBlock(template=t,block_key=f"free_{uuid.uuid4().hex[:8]}",label="Freier Text",content_template="Freier Text",x_mm=20,y_mm=40,width_mm=80,font_size=11,text_color="#FFFFFF",align="left",sort_order=max_order+10)
                db.session.add(block); db.session.commit()
                flash("Freier Deckblatt-Textblock hinzugefügt.","success")
                return redirect(url_for("cover_letter_template_edit",template_id=t.id,tab="cover"))
            if action.startswith("delete_cover_block_"):
                try: block_id=int(action.rsplit("_",1)[1])
                except Exception: block_id=0
                block=CoverLetterCoverBlock.query.filter_by(id=block_id,template_id=t.id).first()
                if block:
                    db.session.delete(block); db.session.commit(); flash("Deckblatt-Textblock gelöscht.","success")
                return redirect(url_for("cover_letter_template_edit",template_id=t.id,tab="cover"))
            if action=="remove_cover":
                if cover.image_filename: safe_remove_cv_media(cover.image_filename)
                cover.image_filename=None; cover.enabled=False
                db.session.commit(); flash("Deckblattgrafik entfernt.","success")
                return redirect(url_for("cover_letter_template_edit",template_id=t.id,tab="cover"))
            t.name=request.form.get("name","").strip() or t.name
            t.description=request.form.get("description","").strip() or None
            color=request.form.get("accent_color","#172033"); t.accent_color=color if re.match(r"^#[0-9A-Fa-f]{6}$",color) else "#172033"
            font=request.form.get("font_family","Helvetica"); t.font_family=font if font in ("Helvetica","Times-Roman","Courier") else "Helvetica"
            try: t.font_scale=max(.8,min(1.2,float(request.form.get("font_scale","1"))))
            except ValueError: t.font_scale=1.0
            t.page_margin_mm=max(12,min(30,request.form.get("page_margin_mm",20,type=int)))
            t.show_logo=request.form.get("show_logo")=="on"; t.is_active=request.form.get("is_active")=="on"; t.sort_order=request.form.get("sort_order",0,type=int)
            save_cover_style_from_form(cover); save_cover_blocks_from_form(t); save_footer_style_from_form(footer); save_template_certification_selection(t)
            try:
                upload=request.files.get("cover_image")
                if upload and upload.filename: save_cover_image_upload(cover,upload)
            except ValueError as exc:
                db.session.rollback(); flash(str(exc),"danger")
                return redirect(url_for("cover_letter_template_edit",template_id=t.id,tab="cover"))
            db.session.commit(); flash("Anschreiben-Template gespeichert.","success")
            return redirect(url_for("cover_letter_template_edit",template_id=t.id,tab=request.form.get("active_tab","general")))
        selected={link.certification_id:link for link in t.footer_certification_links}
        profile=CVProfile.query.first()
        sample_values={
            "applicant_name":(" ".join([profile.first_name or "",profile.last_name or ""]).strip() if profile else "Peter Lange"),
            "applicant_address":(profile.address if profile else "Musterstraße 12") or "Musterstraße 12",
            "applicant_postal_city":(" ".join(x for x in [(profile.postal_code if profile else "12345"),(profile.city if profile else "Musterstadt")] if x)),
            "applicant_phone":(profile.phone if profile else "+49 123 456789") or "+49 123 456789",
            "applicant_email":(profile.email if profile else "peter@example.com") or "peter@example.com",
            "position":"Abteilungsleiter IT","source":"Indeed.com","reference":"B-2026-001",
            "company":"Muster GmbH","department":"Personalabteilung / HR","contact":"Frau Erika Mustermann",
            "company_street":"Musterstraße 1","company_postal_city":"12345 Musterstadt",
            "salary_expectation":"85.000 EUR","application_date":"30.08.2026","letter_date":"30.08.2026",
        }
        block_previews={b.id:render_cover_block(b,sample_values) for b in t.cover_blocks}
        return render_template("cover_letter_template_edit.html",template=t,footer=footer,cover=cover,certifications=certifications,selected=selected,profile=profile,active_tab=request.args.get("tab","general"),block_previews=block_previews)

    @app.route("/applications/<int:application_id>/letters/new",methods=["GET","POST"])
    def application_letter_new(application_id):
        application=Application.query.get_or_404(application_id); templates=CoverLetterTemplate.query.filter_by(is_active=True).order_by(CoverLetterTemplate.sort_order).all()
        if request.method=="POST":
            template=CoverLetterTemplate.query.get_or_404(request.form.get("template_id",type=int)); letter=create_letter_snapshot(application,template); db.session.commit(); log_event(application,"letter",f"Anschreiben-Snapshot erstellt: {letter.title}"); db.session.commit(); flash("Anschreiben wurde aus der Masterversion erstellt.","success"); return redirect(url_for("application_letter_edit",letter_id=letter.id))
        return render_template("application_letter_new.html",application=application,templates=templates,variables=cover_letter_variables(application))

    @app.route("/application-letters/<int:letter_id>/edit",methods=["GET","POST"])
    def application_letter_edit(letter_id):
        letter=ApplicationLetter.query.get_or_404(letter_id); templates=CoverLetterTemplate.query.order_by(CoverLetterTemplate.sort_order).all()
        if request.method=="POST":
            letter.title=request.form.get("title","").strip() or "Anschreiben"; tid=request.form.get("template_id",type=int)
            if tid: letter.template=CoverLetterTemplate.query.get_or_404(tid)
            for field in ["recipient_company","recipient_department","recipient_contact","recipient_street","recipient_postal_code","recipient_city","recipient_country","sender_name","sender_address","sender_postal_code","sender_city","sender_email","sender_phone","subject","salutation","intro_text","body_text","motivation_text","salary_text","closing_text","signoff"]: setattr(letter,field,request.form.get(field,"").strip() or None)
            letter.letter_date=parse_date(request.form.get("letter_date"))
            action=request.form.get("action","save")
            if action=="copy_master_signature":
                safe_remove_cv_media(letter.signature_filename); letter.signature_filename=copy_cv_media(master_signature_filename(),f"letters/{letter.id}","signature"); db.session.commit(); flash("Master-Unterschrift als Snapshot übernommen.","success"); return redirect(url_for("application_letter_edit",letter_id=letter.id))
            if action=="remove_signature": safe_remove_cv_media(letter.signature_filename); letter.signature_filename=None; db.session.commit(); flash("Unterschrift entfernt.","success"); return redirect(url_for("application_letter_edit",letter_id=letter.id))
            db.session.commit(); flash("Anschreiben-Snapshot gespeichert.","success"); return redirect(url_for("application_letter_edit",letter_id=letter.id))
        return render_template("application_letter_edit.html",letter=letter,templates=templates,master_signature=master_signature_filename())

    @app.get("/application-letters/<int:letter_id>/preview")
    def application_letter_preview(letter_id):
        letter=ApplicationLetter.query.get_or_404(letter_id); footer=footer_style_for(letter.template); cover=cover_style_for(letter.template); certs=template_certifications(letter.template)
        cover_values=letter_cover_values(letter)
        cover_blocks=[(block,render_cover_block(block,cover_values)) for block in letter.template.cover_blocks if block.enabled]
        return render_template("application_letter_preview.html",letter=letter,footer=footer,cover=cover,certifications=certs,cover_blocks=cover_blocks,sender_line=letter_sender_line(letter),metrics=letter_render_metrics(letter.template))

    @app.get("/application-letters/<int:letter_id>/pdf")
    def application_letter_pdf(letter_id):
        letter=ApplicationLetter.query.get_or_404(letter_id)
        try:
            pdf=build_application_letter_pdf(letter); filename=f"{secure_filename(letter.application.reference)}_{secure_filename(letter.recipient_company or letter.application.company)}_Anschreiben.pdf"; return send_file(pdf,as_attachment=True,download_name=filename,mimetype="application/pdf")
        except Exception as exc:
            current_app.logger.exception("Anschreiben-PDF fehlgeschlagen für %s",letter_id); flash(f"Anschreiben-PDF konnte nicht erstellt werden: {type(exc).__name__}: {exc}","danger"); return redirect(url_for("application_letter_edit",letter_id=letter_id))

    @app.post("/application-letters/<int:letter_id>/delete")
    def application_letter_delete(letter_id):
        letter=ApplicationLetter.query.get_or_404(letter_id); aid=letter.application_id; media=cv_media_root()/"letters"/str(letter.id); db.session.delete(letter); db.session.commit(); shutil.rmtree(media,ignore_errors=True); flash("Anschreiben-Snapshot gelöscht.","success"); return redirect(url_for("edit_application",application_id=aid))

    @app.get("/applications/<int:application_id>/print")
    def print_application(application_id):
        application=Application.query.get_or_404(application_id)
        return render_template("print_application.html", application=application)

    @app.get("/applications/<int:application_id>/pdf")
    def application_pdf(application_id):
        application=Application.query.get_or_404(application_id)
        pdf=build_application_pdf(application)
        filename=f"{secure_filename(application.reference)}_{secure_filename(application.company)}.pdf"
        return send_file(pdf, as_attachment=True, download_name=filename, mimetype="application/pdf")

    @app.post("/applications/<int:application_id>/attachments")
    def upload_attachment(application_id):
        a=Application.query.get_or_404(application_id); f=request.files.get("file")
        if not f or not f.filename: flash("Keine Datei ausgewählt.","danger"); return redirect(url_for("edit_application",application_id=a.id))
        if not allowed_file(f.filename): flash("Dieser Dateityp ist nicht erlaubt.","danger"); return redirect(url_for("edit_application",application_id=a.id))
        original=secure_filename(f.filename); stored=f"{uuid.uuid4().hex}_{original}"; f.save(Path(app.config["UPLOAD_DIR"])/stored); db.session.add(Attachment(application=a,original_name=original,stored_name=stored)); log_event(a,"attachment",f"Anhang hinzugefügt: {original}"); db.session.commit(); flash("Anhang hochgeladen.","success"); return redirect(url_for("edit_application",application_id=a.id))

    @app.get("/attachments/<int:attachment_id>")
    def download_attachment(attachment_id):
        f=Attachment.query.get_or_404(attachment_id); return send_from_directory(app.config["UPLOAD_DIR"],f.stored_name,as_attachment=True,download_name=f.original_name)

    @app.post("/attachments/<int:attachment_id>/delete")
    def delete_attachment(attachment_id):
        f=Attachment.query.get_or_404(attachment_id); a=f.application; p=Path(app.config["UPLOAD_DIR"])/f.stored_name
        if p.exists(): p.unlink()
        log_event(a,"attachment",f"Anhang gelöscht: {f.original_name}"); db.session.delete(f); db.session.commit(); flash("Anhang gelöscht.","success"); return redirect(url_for("edit_application",application_id=a.id))

    @app.route("/settings/statuses",methods=["GET","POST"])
    def statuses_page():
        ensure_status_classifications()
        if request.method=="POST":
            action=request.form.get("action","add_status")
            if action=="save_classifications":
                for status in get_statuses():
                    classification=status.classification
                    if not classification:
                        classification=StatusClassification(status=status)
                        db.session.add(classification)
                    classification.is_completed=request.form.get(f"completed_{status.id}")=="on"
                db.session.commit()
                flash("Abschluss-Status wurden gespeichert.","success")
                return redirect(url_for("statuses_page"))

            n=request.form.get("name","").strip()
            if not n or Status.query.filter(func.lower(Status.name)==n.lower()).first():
                flash("Status ungültig oder bereits vorhanden.","danger")
            else:
                mx=db.session.query(func.max(Status.sort_order)).scalar()
                status=Status(name=n,sort_order=(mx or 0)+1)
                db.session.add(status)
                db.session.flush()
                db.session.add(StatusClassification(status=status,is_completed=False))
                db.session.commit()
                flash("Status angelegt.","success")
            return redirect(url_for("statuses_page"))
        return render_template("statuses.html")

    @app.post("/settings/statuses/<int:status_id>/rename")
    def rename_status(status_id):
        s=Status.query.get_or_404(status_id); n=request.form.get("name","").strip(); dup=Status.query.filter(Status.id!=s.id,func.lower(Status.name)==n.lower()).first()
        if not n or dup: flash("Status ungültig oder bereits vorhanden.","danger")
        else: s.name=n; db.session.commit(); flash("Status umbenannt.","success")
        return redirect(url_for("statuses_page"))

    @app.post("/settings/statuses/<int:status_id>/delete")
    def delete_status(status_id):
        s=Status.query.get_or_404(status_id)
        if s.applications: flash("Status wird noch verwendet.","danger")
        else: db.session.delete(s); db.session.commit(); flash("Status gelöscht.","success")
        return redirect(url_for("statuses_page"))

    @app.route("/settings/personalize",methods=["GET","POST"])
    def personalize():
        if request.method=="POST":
            action=request.form.get("action","save")
            settings=get_app_settings()
            branding_dir=Path(app.config["UPLOAD_DIR"]) / "branding"
            branding_dir.mkdir(parents=True,exist_ok=True)

            if action=="save_menu":
                save_menu_config_from_form()
                db.session.commit()
                flash("Navigation wurde gespeichert.","success")
                return redirect(url_for("personalize"))

            if action=="reset_menu":
                set_app_setting("menu_config",json.dumps(DEFAULT_MENU_ITEMS,ensure_ascii=False))
                db.session.commit()
                flash("Navigation wurde auf die Standardstruktur zurückgesetzt.","success")
                return redirect(url_for("personalize"))

            if action=="remove_logo":
                old=settings.get("logo_filename","")
                if old:
                    old_path=branding_dir / old
                    if old_path.exists():
                        old_path.unlink()
                set_app_setting("logo_filename","")
                db.session.commit()
                flash("Logo wurde entfernt.","success")
                return redirect(url_for("personalize"))

            app_name=request.form.get("app_name","").strip()
            if not app_name:
                flash("Der Applikationsname darf nicht leer sein.","danger")
                return redirect(url_for("personalize"))

            for key in ("app_name","app_subname","author","copyright_holder","website","footer_text"):
                set_app_setting(key,request.form.get(key,"").strip())

            overview_default_mode=request.form.get("overview_default_mode","active")
            if overview_default_mode not in ("all","active","completed"):
                overview_default_mode="active"
            set_app_setting("overview_default_mode",overview_default_mode)

            logo=request.files.get("logo")
            if logo and logo.filename:
                ext=logo.filename.rsplit(".",1)[1].lower() if "." in logo.filename else ""
                if ext not in LOGO_EXTENSIONS:
                    flash("Logo-Dateityp nicht unterstützt. Erlaubt: PNG, JPG, JPEG, WEBP.","danger")
                    return redirect(url_for("personalize"))
                old=settings.get("logo_filename","")
                if old:
                    old_path=branding_dir / old
                    if old_path.exists():
                        old_path.unlink()
                filename=f"logo_{uuid.uuid4().hex}.{ext}"
                logo.save(branding_dir / filename)
                set_app_setting("logo_filename",filename)

            db.session.commit()
            flash("Personalisierung wurde gespeichert.","success")
            return redirect(url_for("personalize"))
        return render_template("personalize.html",settings=get_app_settings(),menu_items=get_menu_config())

    @app.get("/branding/<path:filename>")
    def branding_logo(filename):
        return send_from_directory(Path(app.config["UPLOAD_DIR"]) / "branding",filename)

    @app.get("/cv-media/<path:filename>")
    def cv_media(filename):
        return send_from_directory(Path(app.config["UPLOAD_DIR"]) / "cv_media",filename)

    @app.get("/applications/print-overview")
    def print_overview():
        query, filters=application_filter_query()
        applications=query.all()
        status_counts={}
        for a in applications:
            name=a.status.name if a.status else "Ohne Status"
            status_counts[name]=status_counts.get(name,0)+1
        return render_template(
            "print_overview.html",
            applications=applications,
            filters=filters,
            status_counts=status_counts
        )

    @app.get("/applications/overview.pdf")
    def overview_pdf():
        query, filters=application_filter_query()
        applications=query.all()
        pdf=build_overview_pdf(applications,filters)
        return send_file(
            pdf,
            as_attachment=True,
            download_name=f"b-v-s_bewerbungsuebersicht_{date.today().isoformat()}.pdf",
            mimetype="application/pdf"
        )

    @app.get("/backups")
    def backups(): return render_template("backups.html")

    @app.post("/backup/download")
    def backup_download():
        """Create a portable ZIP backup without relying on mysqldump binaries."""
        tmp = Path(tempfile.mkdtemp(prefix="bewerbung-backup-"))
        sql = tmp / "database.sql"

        try:
            db_url = os.getenv(
                "DATABASE_URL",
                "mysql+pymysql://bewerbung:bewerbung@db:3306/bewerbungen?charset=utf8mb4"
            )
            parsed = urlparse(db_url.replace("mysql+pymysql://", "mysql://", 1))
            db_host = parsed.hostname or "db"
            db_port = parsed.port or 3306
            db_user = unquote(parsed.username or "bewerbung")
            db_password = unquote(parsed.password or "bewerbung")
            db_name = (parsed.path or "/bewerbungen").lstrip("/") or "bewerbungen"

            import pymysql
            connection = pymysql.connect(
                host=db_host,
                port=db_port,
                user=db_user,
                password=db_password,
                database=db_name,
                charset="utf8mb4",
                autocommit=True
            )

            def qi(name):
                return "`" + str(name).replace("`", "``") + "`"

            with connection.cursor() as cursor, sql.open("w", encoding="utf-8", newline="\n") as out:
                out.write("-- B-V-S SQL Backup\n")
                out.write(f"-- Erstellt: {datetime.now().isoformat(sep=' ', timespec='seconds')}\n\n")
                out.write("SET NAMES utf8mb4;\n")
                out.write("SET FOREIGN_KEY_CHECKS=0;\n")
                out.write("SET UNIQUE_CHECKS=0;\n\n")

                cursor.execute("SHOW FULL TABLES WHERE Table_type = 'BASE TABLE'")
                tables = [row[0] for row in cursor.fetchall()]

                for table in tables:
                    cursor.execute(f"SHOW CREATE TABLE {qi(table)}")
                    create_sql = cursor.fetchone()[1]
                    out.write(f"DROP TABLE IF EXISTS {qi(table)};\n")
                    out.write(create_sql.rstrip(";") + ";\n\n")

                for table in tables:
                    cursor.execute(f"SELECT * FROM {qi(table)}")
                    columns = [desc[0] for desc in cursor.description]
                    column_sql = ",".join(qi(col) for col in columns)

                    while True:
                        rows = cursor.fetchmany(250)
                        if not rows:
                            break

                        values_sql = []
                        for row in rows:
                            escaped = []
                            for value in row:
                                if value is None:
                                    escaped.append("NULL")
                                else:
                                    item = connection.escape(value)
                                    if isinstance(item, bytes):
                                        item = item.decode("utf-8")
                                    escaped.append(str(item))
                            values_sql.append("(" + ",".join(escaped) + ")")

                        out.write(
                            f"INSERT INTO {qi(table)} ({column_sql}) VALUES\n"
                            + ",\n".join(values_sql)
                            + ";\n"
                        )
                    out.write("\n")

                out.write("SET UNIQUE_CHECKS=1;\n")
                out.write("SET FOREIGN_KEY_CHECKS=1;\n")

            connection.close()

        except Exception as exc:
            shutil.rmtree(tmp, ignore_errors=True)
            flash(f"Backup konnte nicht erstellt werden: {type(exc).__name__}: {exc}", "danger")
            return redirect(url_for("backups"))

        archive = tmp / f"b-v-s_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        with zipfile.ZipFile(archive, "w", zipfile.ZIP_DEFLATED) as z:
            z.write(sql, "database.sql")

            upload_dir = Path(app.config["UPLOAD_DIR"])
            if upload_dir.exists():
                for path in upload_dir.rglob("*"):
                    if path.is_file():
                        z.write(path, Path("attachments") / path.relative_to(upload_dir))

            restore = """B-V-S – Wiederherstellung

1. Anwendung stoppen:
   docker compose down

2. Anwendung/Datenbank starten:
   docker compose up -d

3. Datenbank wiederherstellen:
   docker compose exec -T db mysql -u bewerbung -pbewerbung bewerbungen < database.sql

4. Den Inhalt des Ordners attachments/ nach ./data/attachments/ kopieren.

Hinweis:
Die Datenbank-Zugangsdaten im Restore-Befehl müssen zu deiner docker-compose.yml passen.
"""
            z.writestr("RESTORE.txt", restore)

        return send_file(
            archive,
            as_attachment=True,
            download_name=archive.name,
            mimetype="application/zip"
        )

    @app.errorhandler(413)
    def too_large(_): flash("Datei zu groß (max. 20 MB).","danger"); return redirect(request.referrer or url_for("index"))
    return app

app=create_app()
